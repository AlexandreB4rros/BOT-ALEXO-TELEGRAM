# Importações de bibliotecas padrão do Python
import os
import sys
import json
import shutil
import zipfile
import logging
import xml.etree.ElementTree as ET
import glob
from functools import wraps
import io
import secrets
from datetime import time as dt_time
import pytz
from pathlib import Path
import re

# Importações de bibliotecas de terceiros (instaladas via pip)
import aiohttp
import openpyxl
from openpyxl import Workbook
import warnings
import simplekml
import aiomysql
import pandas as pd
import matplotlib.pyplot as plt
import contextily as cx
from dotenv import load_dotenv
from telegram import Update
from telegram.ext import (ApplicationBuilder, CommandHandler, ContextTypes, MessageHandler, filters, ConversationHandler, JobQueue, TypeHandler)  # Componentes para construir o bot
from telegram.error import NetworkError, Forbidden
from Scripts_Alexo import selecionar_token, __version__
import asyncio
import aiofiles
from html import escape
from telegram.constants import ParseMode
import httpx

caminho_env = Path(__file__).parent / ".env"

# Carrega as variáveis de ambiente do arquivo .env localizado no mesmo diretório do script
load_dotenv(dotenv_path=caminho_env)

warnings.filterwarnings("ignore", category=UserWarning)

class IgnoreAttributeErrorFilter(logging.Filter):
    # Este filtro impede que mensagens de log contendo "AttributeError" sejam exibidas.
    def filter(self, record):
        return "AttributeError" not in record.getMessage()

async def send_log_to_telegram(message: str):
    url = f'https://api.telegram.org/bot{BOT_TOKEN}/sendMessage'
    
    log_escapado = escape(str(message))
    
    payload = {
        'chat_id': TELEGRAM_GROUP_ID,
        'text': f"<b>[LOG]</b>\n<pre><code>{log_escapado}</code></pre>",
        'parse_mode': ParseMode.HTML  
    }
    
    try:
        timeout = aiohttp.ClientTimeout(total=5)
        async with aiohttp.ClientSession(timeout=timeout) as session:
            async with session.post(url, json=payload) as response:
                if response.status != 200:
                    response_text = await response.text()
                    print(f"LOGGING FALLBACK (API ERROR): Status {response.status} ao enviar log. Resposta: {response_text}")

    except (aiohttp.ClientError, asyncio.TimeoutError) as e:
        print(f"LOGGING FALLBACK (NETWORK ERROR): {e}")
    except Exception as e:
        print(f"LOGGING FALLBACK (UNEXPECTED ERROR): {e}")



logger = logging.getLogger()
logger.setLevel(logging.INFO)

console_handler = logging.StreamHandler()
console_handler.setLevel(logging.INFO)
console_formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
console_handler.setFormatter(console_formatter)
console_handler.addFilter(IgnoreAttributeErrorFilter())


logger.addHandler(console_handler)


# --- Metadados e Constantes Globais ---

# Define os metadados do script
__author__ = "Alexandre B, J. Ayrton"
__credits__ = "Anderson, Josimar"

ROOT_DIR = Path(__file__).parent
# Constrói o caminho absoluto para o ficheiro JSON
FILENAME_WEBHOOK = ROOT_DIR / "WebHook.json"

# Limita o traceback do Python para não exibir rastreamentos detalhados de erro
sys.tracebacklimit = 0

DBUG = 2


# --- Inicialização do Token do Bot ---

# Seleciona o token do bot conforme o modo de debug
try:
    # Tenta obter o token do bot chamando a função customizada 'selecionar_token'
    BOT_TOKEN = selecionar_token(DBUG)
    DB_DATABASE = selecionar_token(DBUG)
except ValueError as e:
    logger.error(f"Erro: {e}")

# --- Constantes de Mensagens ---

# Define um 'set' com os valores válidos para splitters.
SPLITTERS_VALIDOS = {"1/16", "1/8", "1/4"}

TELEGRAM_GROUP_ID = "-1002292627707" #GRUPO DE LOGS

# Mensagens de erro pré-definidas para serem usadas nas respostas do bot.
ErroE101 = "❌ Atenção, excesso de argumentos. Verifique o comando informado e tente novamente!"
ErroP101 = "❌ Atenção, 'POP' não informado!"
ErroP102 = "❌ Atenção, 'POP' não existe na lista de templates. Verifique se foi informado corretamente ou notifique a equipe interna."
ErroF101 = "❌ Atenção, 'FSAN/SN' não informado para a consulta. Verifique o comando e tente novamente!"
ErroF102 = "❌ Atenção, O formato do campo 'FSAN/SN' está incorreto!"
ErroS101 = "❌ Atenção, 'SPLITTER' não informado. Verifique o comando e tente novamente!"
ErroN101 = "❌ Atenção, 'OLT/SLOT/PON' não informado. Verifique o comando e tente novamente!"
ErroN102 = "❌ Atenção, 'OLT/SLOT/PON' contém mais de duas '/'. Verifique o comando e tente novamente!"
ErroC101 = "❌ Atenção, verifique se a 'CTO' informada está correta e tente novamente."

# --- Manipulador de Erros ---

# Função para lidar com todos os erros capturados pela biblioteca python-telegram-bot.
async def error_handler(update: object, context: ContextTypes.DEFAULT_TYPE) -> None:
    err = context.error
    
    # Verifica se o erro é uma falha de conexão de rede com a API do Telegram.
    if isinstance(err, NetworkError):
        # Verifica se o bot já não está marcado como desconectado para evitar notificações repetidas.
        if not context.bot_data.get('is_disconnected', False):
            # Registra um aviso sobre a perda de conexão.
            logger.warning("Conexão com o Telegram perdida. Marcando como desconectado.")
            context.bot_data['is_disconnected'] = True
            mensagem_para_admin = (f"🚨 ALERTA DE CONEXÃO 🚨\n\nO bot perdeu a conexão com o Telegram.\n\nErro: {err}")
            # Notifica os administradores sobre a queda.
            await notificar_admins(context, mensagem_para_admin)
        return
        
    # Para qualquer outro tipo de erro, registra a exceção completa para depuração.
    logger.error(f"Exceção capturada pelo handler global: {err}", exc_info=True)
    mensagem_generica = f"Ocorreu uma exceção não tratada no bot: {err}"
    await notificar_admins(context, mensagem_generica)


# --- Verificador de Reconexão ---

async def check_reconnection(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    is_disconnected_flag = context.bot_data.get('is_disconnected', False)
    
    # Se a flag estiver True, significa que o bot estava desconectado e agora recebeu uma atualização.
    if is_disconnected_flag:
        logger.info("Conexão com o Telegram reestabelecida. Enviando notificação para admins.")
        
        # Reseta a flag para False, indicando que o bot está online.
        context.bot_data['is_disconnected'] = False 
        mensagem_online = "✅ *CONEXÃO REESTABELECIDA* ✅\n\nO bot está online e operando normalmente."
        # Notifica os administradores que o bot está de volta.
        await notificar_admins(context, mensagem_online)

# Função para centralizar e padronizar a criação de conexões com o banco de dados.
async def criar_conexao_db():
    """Cria e retorna uma conexão assíncrona com o banco de dados."""
    try:
        # Carrega as credenciais do banco de dados a partir de variáveis de ambiente.
        return await aiomysql.connect(
            host=os.getenv("DB_HOST"),
            user=os.getenv("DB_USER"),
            password=os.getenv("DB_PASSWORD"),
            db=DB_DATABASE,
            connect_timeout=5,
            autocommit=True # Autocommit para simplificar operações
        )
    except Exception as e:
        logger.error(f"Falha ao criar conexão com o DB: {e}")
        return None


# --- Sistema de Notificação de Administradores com Fallback ---

# Tarefa agendada para periodicamente salvar uma lista de admins em um arquivo JSON.
async def atualizar_admins_fallback(context: ContextTypes.DEFAULT_TYPE):
    logger.info("Executando tarefa agendada: Atualizando lista de administradores de fallback...")
    conexao_db = await criar_conexao_db()
    if not conexao_db:
        logger.error("Job 'atualizar_admins_fallback': Não foi possível conectar ao DB.")
        return
    try:
        async with conexao_db.cursor(aiomysql.DictCursor) as cursor:
            # Query SQL para selecionar o ID do Telegram de todos os usuários com o cargo 'Administrador'.
            query = "SELECT u.id_telegram FROM usuarios u JOIN cargos c ON u.cargo_id = c.id WHERE c.nome_cargo = 'Administrador'"
            await cursor.execute(query)
            admins = await cursor.fetchall()
            # Extrai apenas os IDs da lista de dicionários.
            admin_ids = [admin['id_telegram'] for admin in admins]
        # Abre (ou cria) o arquivo 'admins_fallback.json' para escrita de forma assíncrona.
        async with aiofiles.open("admins_fallback.json", "w", encoding="utf-8") as f:
            # Salva a lista de IDs de admin no arquivo JSON.
            await f.write(json.dumps({"admin_ids": admin_ids}))
        logger.info(f"Lista de fallback de administradores atualizada com sucesso. {len(admin_ids)} admin(s) salvo(s).")
    except Exception as e:
        logger.error(f"Job 'atualizar_admins_fallback': Falha ao executar a tarefa. Erro: {e}")
    finally:
        if conexao_db:
            conexao_db.close()


# Função principal para notificar administradores.
# A primeira tentativa é sempre buscar a lista de admins direto do banco de dados.
async def notificar_admins(context: ContextTypes.DEFAULT_TYPE, mensagem_erro: str):
    """
    Busca a lista de administradores no DB e os notifica sobre um erro,
    ignorando de forma segura os utilizadores que bloquearam o bot.
    """
    conexao_db = None
    try:
        conexao_db = await criar_conexao_db()
        if not conexao_db:
            raise ConnectionError("Falha ao obter conexão com o DB para notificação.")
        
        async with conexao_db.cursor(aiomysql.DictCursor) as cursor:
            query = "SELECT u.id_telegram FROM usuarios u JOIN cargos c ON u.cargo_id = c.id WHERE c.nome_cargo = 'Administrador'"
            await cursor.execute(query)
            admins = await cursor.fetchall()

        if not admins:
            logger.warning("Nenhum administrador encontrado no banco de dados para notificar.")
            return
            
        erro_escapado = escape(str(mensagem_erro))
        mensagem_formatada = (
            f"<b>🚨 ALERTA DE ERRO 🚨</b>\n\n"
            f"Ocorreu a seguinte falha no bot:\n"
            f"<pre><code>{erro_escapado}</code></pre>"
        )

        # --- LÓGICA DE ENVIO MELHORADA ---
        tasks = []
        for admin in admins:
            admin_id = admin['id_telegram']
            try:
                # Cria a tarefa de envio para cada admin
                task = context.bot.send_message(
                    chat_id=admin_id,
                    text=mensagem_formatada,
                    parse_mode=ParseMode.HTML
                )
                tasks.append(task)
            except Forbidden:
                # Se o bot for bloqueado, regista um aviso e continua para o próximo
                logger.warning(f"Não foi possível notificar o admin {admin_id}. O bot foi bloqueado ou não foi iniciado.")
            except Exception as e:
                logger.error(f"Erro inesperado ao tentar enviar mensagem para o admin {admin_id}: {e}")

        # Envia todas as mensagens válidas em paralelo
        if tasks:
            # Usamos return_exceptions=True para garantir que, se um envio falhar por outro motivo,
            # os outros não sejam interrompidos.
            await asyncio.gather(*tasks, return_exceptions=True)

    except Exception as db_err:
        logger.error(f"Falha ao notificar admins (DB), acionando fallback. Erro: {db_err}")
        await notificar_admins_fallback(context, mensagem_erro)
    finally:
        if conexao_db:
            conexao_db.close()

# Função de notificação de fallback, usada quando o banco de dados está inacessível.
async def notificar_admins_fallback(context: ContextTypes.DEFAULT_TYPE, mensagem_erro: str):
    """
    Notifica os admins usando uma lista de um ficheiro JSON, usada quando o DB falha.
    Também ignora de forma segura os utilizadores que bloquearam o bot.
    """
    logger.warning("Acionando modo de notificação de fallback (lendo do arquivo JSON).")
    try:
        # Abre o arquivo JSON que contém a lista de admins salva.
        async with aiofiles.open("admins_fallback.json", "r", encoding="utf-8") as f:
            dados = json.loads(await f.read())
            admin_ids = dados.get("admin_ids", [])
            
        if not admin_ids:
            logger.error("O arquivo de fallback de administradores está vazio ou não foi encontrado. Ninguém foi notificado.")
            return

        # Formata a mensagem de forma segura com HTML
        erro_escapado = escape(str(mensagem_erro))
        mensagem_formatada = (
            f"<b>🚨 ALERTA DE ERRO (Notificação de Fallback) 🚨</b>\n\n"
            f"A notificação primária falhou. O erro original reportado foi:\n\n"
            f"<pre><code>{erro_escapado}</code></pre>"
        )

        tasks = []
        for admin_id in admin_ids:
            try:
                task = context.bot.send_message(
                    chat_id=admin_id, 
                    text=mensagem_formatada, 
                    parse_mode=ParseMode.HTML
                )
                tasks.append(task)
            except Forbidden:
                logger.warning(f"Fallback: Não foi possível notificar o admin {admin_id}. O bot foi bloqueado ou não foi iniciado.")
            except Exception as e:
                logger.error(f"Erro inesperado na função de notificação de fallback para o admin {admin_id}: {e}")

        if tasks:
            await asyncio.gather(*tasks, return_exceptions=True)
            
    except FileNotFoundError:
        logger.error("Arquivo 'admins_fallback.json' não encontrado. Não foi possível executar a notificação de fallback.")
    except Exception as e:
        logger.error(f"Erro crítico inesperado na função de notificação de fallback: {e}")


# --- Decorator de Verificação de Permissão ---


def check_permission(func):
    """
    Decorador para verificar se:
    1. O usuario está ativo.
    2. O usuario tem permissão para o comando.
    3. Atualiza a data de ultima interação do usuario.
    """
    @wraps(func)
    async def wrapper(update: Update, context: ContextTypes.DEFAULT_TYPE, *args, **kwargs):
        message = update.message or update.edited_message
        user = update.effective_user

        if not message or not user:
            logger.warning("Recebido um update sem mensagem ou usuario efetivo. Ignorando.")
            return

        chat = message.chat
        user_id = user.id
        command_name = func.__name__
        conexao_db = None
        
        try:
            conexao_db = await criar_conexao_db()
            if not conexao_db:
                raise ConnectionError("DB indisponível para checar permissão.")
                
            async with conexao_db.cursor(aiomysql.DictCursor) as cursor:
                # 1. Verifica se o utilizador está ativo E se tem permissão
                query = """
                    SELECT u.esta_ativo 
                    FROM usuarios u
                    JOIN permissoes p ON u.cargo_id = p.cargo_id 
                    JOIN comandos cmd ON p.comando_id = cmd.id 
                    WHERE u.id_telegram = %s AND cmd.nome_comando = %s 
                    LIMIT 1;
                """
                await cursor.execute(query, (user_id, command_name))
                resultado = await cursor.fetchone()

                if resultado:
                    # Verifica se a conta está ativa
                    if not resultado['esta_ativo']:
                        await chat.send_message("❌ A sua conta está inativa por falta de uso. Por favor, Entre em contato com seu supervisor.")
                        return

                    # 2. Atualiza a data de última interação
                    await cursor.execute(
                        "UPDATE usuarios SET ultima_interacao = NOW() WHERE id_telegram = %s",
                        (user_id,)
                    )
                    
                    # 3. Executa o comando solicitado
                    await func(update, context, *args, **kwargs)
                else:
                    await chat.send_message("❌ Você não tem permissão para usar este comando.")
                    
        except Exception as err:
            error_message = f"Erro na verificação de permissão para o comando /{command_name}: {err}"
            logger.error(error_message, exc_info=True)
            await notificar_admins(context, error_message)
            await chat.send_message("⚠️ Ocorreu um erro ao verificar as suas permissões. A equipe de administração foi notificada.")

        finally:
            if conexao_db:
                conexao_db.close()
                
    return wrapper

# --- Comandos ---

# Comando /cadastrar.
@check_permission
async def cadastrar(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    message = update.message or update.edited_message
    conexao_db = None
    try:
        conexao_db = await criar_conexao_db()
        if not conexao_db:
            raise ConnectionError("DB indisponível.")
        
        # Se o comando for executado sem argumentos, lista os cargos.
        if len(context.args) != 1:
            async with conexao_db.cursor(aiomysql.DictCursor) as cursor:
                await cursor.execute("SELECT nome_cargo FROM cargos ORDER BY nome_cargo;")
                resultados = await cursor.fetchall()

            lista_cargos_segura = "\n".join(
                [f"  - {escape(item['nome_cargo'])}" for item in resultados]
            ) if resultados else "Nenhum cargo encontrado."

            mensagem_ajuda = (
                f"Uso: <code>/cadastrar &lt;CARGO&gt;</code>\n\n"
                f"<b>Cargos disponíveis:</b>\n{lista_cargos_segura}"
            )
            
            await message.reply_text(mensagem_ajuda, parse_mode=ParseMode.HTML)
            return

        # Se o comando tiver um argumento, inicia a geração do convite.
        cargo_solicitado = context.args[0].capitalize()
        async with conexao_db.cursor(aiomysql.DictCursor) as cursor:
            await cursor.execute("SELECT id FROM cargos WHERE nome_cargo = %s", (cargo_solicitado,))
            resultado_cargo = await cursor.fetchone()
            if not resultado_cargo:
                await message.reply_text(f"❌ Cargo '{escape(cargo_solicitado)}' inválido. Verifique os cargos com /cadastrar.")
                return

            cargo_id = resultado_cargo['id']
            hash_convite = secrets.token_hex(16)
            query_insert_invite = "INSERT INTO cadastros_pendentes (hash_convite, cargo_id, admin_id) VALUES (%s, %s, %s)"
            await cursor.execute(query_insert_invite, (hash_convite, cargo_id, user.id))
            
            bot_info = await context.bot.get_me()
            bot_username = bot_info.username
            cargo = escape(str(cargo_solicitado))

            mensagem = (
                f"✅ Convite de cadastro gerado com sucesso!\n\n"
                f"<b>Cargo:</b> {cargo}\n\n"
                f"Peça para o novo usuário contatar o bot @{bot_username} e enviar o seguinte comando:\n\n"
                f"(Clique no texto abaixo para copiar 👇)\n"
                f"<code>/novo_usuario {hash_convite}</code>")
            
            await message.reply_text(mensagem, parse_mode=ParseMode.HTML)
            logger.info(f"Admin {user.id} gerou um convite para o cargo {cargo_solicitado} (ID: {cargo_id})")

    except Exception as e:
        await message.reply_text("Ocorreu um erro ao processar o cadastro.")
        logger.error(f"Erro no comando /cadastrar: {e}", exc_info=True)
    finally:
        if conexao_db:
            conexao_db.close()


# --- Fluxo de Conversa para Cadastro de Novo Usuário ---

# Define os "estados" da conversa.
VERIFICAR_HASH, RECEBER_MATRICULA, RECEBER_NOME = range(3)

# Comando /novo_usuario <hash>.
async def novo_usuario(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    message = update.message or update.edited_message
    if not message: return ConversationHandler.END

    if not context.args or len(context.args) != 1:
        await message.reply_text("Uso: /novo_usuario <código_de_convite>")
        return ConversationHandler.END
    
    hash_convite = context.args[0]
    conexao_db = None
    try:
        conexao_db = await criar_conexao_db()
        if not conexao_db: raise ConnectionError("DB indisponível")
        
        async with conexao_db.cursor(aiomysql.DictCursor) as cursor:
            query = "SELECT cp.cargo_id, c.nome_cargo FROM cadastros_pendentes cp JOIN cargos c ON cp.cargo_id = c.id WHERE cp.hash_convite = %s"
            await cursor.execute(query, (hash_convite,))
            resultado = await cursor.fetchone()
        
        if not resultado:
            user = update.effective_user
            username_text = f"@{user.username}" if user.username else "Não definido"
            
            mensagem_para_admins = (
                f"🚨 <b>Tentativa de Cadastro com Convite Inválido</b> 🚨\n\n"
                f"O utilizador abaixo tentou se registar com um código inválido ou já utilizado:\n\n"
                f"👤 <b>Nome:</b> {escape(user.full_name)}\n"
                f"🆔 <b>ID do Telegram:</b> <code>{user.id}</code>\n"
                f"🔗 <b>Username:</b> {escape(username_text)}\n\n"
                f"O código informado foi:\n<code>{escape(hash_convite)}</code>"
            )
            await notificar_admins(context, mensagem_para_admins)
            await message.reply_text("❌ Código de convite inválido ou já utilizado.")
            return ConversationHandler.END

        context.user_data['cadastro_cargo_id'] = resultado['cargo_id']
        context.user_data['cadastro_cargo_nome'] = resultado['nome_cargo']
        context.user_data['cadastro_hash'] = hash_convite
        
        cargo = escape(resultado['nome_cargo'])
        mensagem_bem_vindo = f"✅ Convite válido para o cargo de <b>{cargo}</b>!\n\n"
        mensagem_bem_vindo += f"Olá, <b>{escape(update.effective_user.full_name)}</b>! Bem-vindo(a) ao processo de cadastro.\n\n"
        mensagem_bem_vindo += f"Para concluir o cadastro, precisamos de algumas informações suas.\n\n"
        mensagem_bem_vindo += f"Primeiro, informe a sua matrícula, ela deve ser a mesma constando na plataforma LG ou no crachá de funcionário (Lembre-se que a matrícula é um número de até 6 dígitos).\n"
        mensagem_bem_vindo += f"Caso tenha duvidas acesse: https://login.lg.com.br/login/desktop e faça o login utilizando os dados informados para acesso a plataforma.\n"
        mensagem_bem_vindo += f"Na pagina principal localize o icone do perfil e nele constará a matrícula ao lado do seu nome.\n"
        mensagem_bem_vindo += f"Se precisar de ajuda, use o comando: <code>'/cancelar_cadastro'</code> a qualquer momento para cancelar o processo (Você pode reiniciar o processo de cadastro a qualquer momento, com o mesmo convite recebido).\n\n"
        mensagem_bem_vindo += f"Vamos começar!\n\n"
        mensagem_bem_vindo += f"Por favor, informe sua matrícula:"

        await message.reply_text(mensagem_bem_vindo, parse_mode=ParseMode.HTML)
        return RECEBER_MATRICULA
        
    except Exception as e:
        logger.error(f"Erro em /novo_usuario: {e}", exc_info=True)
        await message.reply_text("Ocorreu um erro ao verificar o convite.")
        return ConversationHandler.END
    finally:
        if conexao_db: conexao_db.close()


async def receber_matricula(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    message = update.message or update.edited_message
    if not message:
        return ConversationHandler.END
    matricula = message.text
    if not matricula.isdigit() or len(matricula) > 6:
        await message.reply_text("❌ Matrícula inválida. Por favor, informe apenas números com no máximo 6 dígitos.")
        # Permanece no mesmo estado para aguardar uma nova tentativa
        return RECEBER_MATRICULA
    
    # Armazena a matrícula recebida.
    context.user_data['cadastro_matricula'] = matricula
    logger.info(f"Usuário {update.effective_user.id} informou a matrícula: {matricula}")
    
    await update.message.reply_text("Obrigado! Agora, por favor, informe seu nome completo:")
    return RECEBER_NOME

async def receber_nome(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    message = update.message or update.edited_message
    if not message:
        return ConversationHandler.END
    user = update.effective_user
    nome_completo = message.text
    
    # Recupera os dados da conversa armazenados em user_data.
    cargo_id = context.user_data.get('cadastro_cargo_id')
    cargo_nome = context.user_data.get('cadastro_cargo_nome')
    hash_convite = context.user_data.get('cadastro_hash')
    matricula = context.user_data.get('cadastro_matricula')
    conexao_db = None
    try:
        conexao_db = await criar_conexao_db()
        if not conexao_db:
            await update.message.reply_text("⚠️ Erro de conexão com o banco de dados para finalizar.")
            return ConversationHandler.END

        async with conexao_db.cursor(aiomysql.DictCursor) as cursor:
            # Query para inserir um novo usuário ou atualizar um existente caso a id do telegram seja igual.
            query_insert_update = """
                INSERT INTO usuarios (id_telegram, nome_usuario, matricula, nome_completo, cargo_id, hash_convite)
                VALUES (%s, %s, %s, %s, %s, %s)
                AS new
                ON DUPLICATE KEY UPDATE 
                    nome_usuario = new.nome_usuario, 
                    matricula = new.matricula, 
                    nome_completo = new.nome_completo, 
                    cargo_id = new.cargo_id,
                    hash_convite = new.hash_convite
            """
            
            await cursor.execute(query_insert_update, (user.id, user.full_name, matricula, nome_completo, cargo_id, hash_convite))
            
            # Deleta o convite para invalidá-lo.
            await cursor.execute("DELETE FROM cadastros_pendentes WHERE hash_convite = %s", (hash_convite,))
        
        await update.message.reply_text(f"✅ Cadastro concluído com sucesso! Bem-vindo(a), {nome_completo}!")
        logger.info(f"Novo usuário cadastrado/atualizado: {user.id}, Nome: {nome_completo}, Matrícula: {matricula}, Cargo: {cargo_nome}")
        return ConversationHandler.END

    except aiomysql.IntegrityError:
        await update.message.reply_text(f"❌ Falha no cadastro. A matrícula informada {matricula} já está em uso por outro usuário. Acesse https://login.lg.com.br/login/desktop e verifique se a matrícula enviada está correta.")
        return ConversationHandler.END
    except Exception as err:
        logger.error(f"Erro de DB na finalização do cadastro: {err}", exc_info=True)
        error_message = f"Erro ao finalizar o cadastro do usuário {user.id}: {err}"
        await notificar_admins(context, error_message)
        await update.message.reply_text("⚠️ Ocorreu um erro ao finalizar seu cadastro. A equipe de administração foi notificada.")
        return ConversationHandler.END
    finally:
        if conexao_db:
            conexao_db.close()
        # Limpa todos os dados temporários da conversa, independentemente do resultado.
        for key in ['cadastro_cargo_id', 'cadastro_cargo_nome', 'cadastro_hash', 'cadastro_matricula']:
            context.user_data.pop(key, None)

# Comando /cancelar_cadastro.
async def cancelar_cadastro(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    # Percorre uma lista de chaves e remove cada uma do 'user_data'.
    for key in ['cadastro_cargo_id', 'cadastro_cargo_nome', 'cadastro_hash', 'cadastro_matricula']:
        context.user_data.pop(key, None) # O 'None' evita erros se a chave não existir.
    # Envia uma mensagem de confirmação para o usuário.
    await update.message.reply_text("Cadastro cancelado.")
    return ConversationHandler.END

# --- Funções de Apoio e Comandos ---

# Comando /Listar_admins.
@check_permission
async def listar_admins(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    message = update.message or update.edited_message
    if not message: return

    chat_id = message.chat.id
    try:
        administradores = await context.bot.get_chat_administrators(chat_id)
        if not administradores:
            await message.reply_text("Não foi possível encontrar administradores neste grupo.")
            return
        
        lista_texto = ["<b>Administradores do Grupo:</b>"]
        for admin in administradores:
            user = admin.user
            lista_texto.append(f"- {escape(user.full_name)} (ID: <code>{user.id}</code>)")
        
        mensagem_final = "\n".join(lista_texto)
        await message.reply_text(mensagem_final, parse_mode=ParseMode.HTML)
    except Exception as e:
        logger.error(f"Erro ao listar administradores: {e}")
        await message.reply_text("Ocorreu um erro ao buscar a lista de administradores. Verifique se o bot tem permissão para isso.")



# --- Funções de Geolocalização e Mapa ---

# Função para buscar CTOs (Caixas de Terminação Óptica) em um raio de uma dada coordenada.
async def buscar_ctos_proximas(lat, lon):
    conexao_db = None
    try:
        # Define o raio de busca em quilômetros.
        raio_km = 0.15
        # Query SQL que utiliza a fórmula de Haversine para calcular a distância
        # entre a coordenada fornecida e cada CTO no banco de dados.
        # 6371 é o raio aproximado da Terra em km.
        query_haversine = """
            SELECT cto, latitude, longitude,
                   (6371 * ACOS(
                       COS(RADIANS(%s)) * COS(RADIANS(latitude)) *
                       COS(RADIANS(longitude) - RADIANS(%s)) +
                       SIN(RADIANS(%s)) * SIN(RADIANS(latitude))
                   )) AS distancia
            FROM ctos
            HAVING distancia <= %s
            ORDER BY distancia;
        """
        conexao_db = await criar_conexao_db()
        if not conexao_db:
            logger.error("Não foi possível conectar ao DB para buscar CTOs.")
            return None
        async with conexao_db.cursor(aiomysql.DictCursor) as cursor:
            # Executa a query passando a latitude, longitude e o raio como parâmetros.
            await cursor.execute(query_haversine, (lat, lon, lat, raio_km))
            ctos_encontradas = await cursor.fetchall()
        return ctos_encontradas
    except Exception as err:
        logger.error(f"Erro de SQL ao buscar CTOs próximas: {err}")
        return None # Retorna None se ocorrer um erro de SQL.
    finally:
        # Garante que a conexão seja fechada.
        if conexao_db:
            conexao_db.close()


# Função para gerar uma imagem de mapa com a localização do usuário e as CTOs próximas.

def mapa_ctos(user_lat, user_lon, ctos_encontradas):
    # Cria DataFrames com os dados das CTOs e do usuário para facilitar a manipulação.
    df_ctos = pd.DataFrame(ctos_encontradas)
    df_user = pd.DataFrame([{'latitude': user_lat, 'longitude': user_lon}])
    # Combina todos os pontos para calcular a área total que o mapa deve cobrir.
    all_points = pd.concat([df_ctos[['latitude', 'longitude']], df_user[['latitude', 'longitude']]])
    # Garante que as coordenadas sejam numéricas, descartando valores inválidos.
    all_points['latitude'] = pd.to_numeric(all_points['latitude'], errors='coerce')
    all_points['longitude'] = pd.to_numeric(all_points['longitude'], errors='coerce')
    all_points.dropna(inplace=True)
    # Calcula os limites do mapa (zoom) com uma margem para que os pontos não fiquem na borda.
    lat_range = all_points.latitude.max() - all_points.latitude.min()
    lon_range = all_points.longitude.max() - all_points.longitude.min()
    padding_percentage = 0.10
    lat_buffer = max(lat_range * padding_percentage, 0.002) # Margem mínima.
    lon_buffer = max(lon_range * padding_percentage, 0.002)
    min_lat, max_lat = all_points.latitude.min() - lat_buffer, all_points.latitude.max() + lat_buffer
    min_lon, max_lon = all_points.longitude.min() - lon_buffer, all_points.longitude.max() + lon_buffer
    # --- Criação do Gráfico (Mapa) ---
    fig, ax = plt.subplots(figsize=(8, 8))
    ax.set_axis_off()
    ax.set_xlim(min_lon, max_lon)
    ax.set_ylim(min_lat, max_lat)
    # Plota as CTOs e a localização do usuário.
    ax.scatter(df_ctos.longitude, df_ctos.latitude, c='red', s=60, zorder=2, label='CTOs Próximas', edgecolors='black')
    for _, row in df_ctos.iterrows():
        ax.text(row.longitude, row.latitude + (lat_buffer * 0.04), f'{row.cto}', fontsize=10, zorder=3, ha='center', fontweight='bold')
    ax.scatter(user_lon, user_lat, c='blue', s=250, marker='o', zorder=4, label='Sua Localização', edgecolors='white')
    # Adiciona um mapa base ao fundo do gráfico.
    cx.add_basemap(ax, crs='EPSG:4326', source=cx.providers.OpenStreetMap.Mapnik)
    # Salva a imagem em memória.
    buf = io.BytesIO()
    plt.savefig(buf, format='png', bbox_inches='tight', pad_inches=0)
    plt.close(fig)
    buf.seek(0)
    return buf

async def criar_mapa_ctos(user_lat, user_lon, ctos_encontradas):
    return await asyncio.to_thread(mapa_ctos, user_lat, user_lon, ctos_encontradas)

# --- Comando para Solicitar Localização ---

# Comando /ctos
@check_permission
async def ctos(update: Update, context: ContextTypes.DEFAULT_TYPE):
    # Solicita ao usuário que envie sua localização e define uma "flag"
    # indicando que o bot está aguardando uma mensagem de localização para buscar CTOs.
    context.user_data['waiting_for_ctos_location'] = True
    await update.message.reply_text("📍 Por favor, envie sua localização para que eu possa encontrar as CTOs mais próximas.")


# --- Manipulador de Localização Unificado ---

# Este handler recebe TODAS as mensagens de localização e decide o que fazer com base nas flags definidas em 'user_data'.
async def unified_location_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    message = update.message or update.edited_message
    
    latitude, longitude = None, None

    # --- Extração Unificada de Coordenadas ---
    if message and message.location:
        latitude, longitude = message.location.latitude, message.location.longitude
    elif message and message.text:
        # Lógica de extração de texto (com todas as tentativas)
        direct_match = re.search(r"(-?\d+\.\d+)[, ]+(-?\d+\.\d+)", message.text)
        if direct_match:
            latitude, longitude = map(float, direct_match.groups())
        else:
            url_match = re.search(r"https?://\S+", message.text)
            if url_match:
                url = url_match.group(0)
                try:
                    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'}
                    async with httpx.AsyncClient(follow_redirects=True) as client:
                        response = await client.get(url, headers=headers, timeout=10)
                    
                    final_url = str(response.url)
                    
                    match_url = (re.search(r"/@(-?\d+\.\d+),(-?\d+\.\d+)", final_url) or
                                 re.search(r"!3d(-?\d+\.\d+)!4d(-?\d+\.\d+)", final_url) or
                                 re.search(r"[?&](?:q|ll|query)=(-?\d+\.\d+),(-?\d+\.\d+)", final_url))
                    
                    if match_url:
                        latitude, longitude = map(float, match_url.groups())
                    else:
                        page_content = response.text
                        match_html = re.search(r'\[null,null,(-?\d+\.\d+),(-?\d+\.\d+)\]', page_content)
                        if match_html:
                            latitude, longitude = map(float, match_html.groups())
                except Exception as e:
                    logger.error(f"Falha crítica ao processar URL '{url}': {e}", exc_info=True)

    # --- Validação e Saída se não houver coordenadas ---
    if latitude is None:
        if context.user_data.get('waiting_for_ctos_location') or context.user_data.get('waiting_for_location'):
            await message.reply_text("Localização não reconhecida. Por favor, envie uma localização válida ou um link do mapa.")
        return
    
    user = update.effective_user

    # --- Lógica de Rotas ---
    
    # Rota 1: Buscar CTOs (aceita link ou nativa)
    if context.user_data.pop('waiting_for_ctos_location', False):
        await message.reply_text("Buscando CTOs em um raio de 150 metros... 📡")
        ctos_encontradas = await buscar_ctos_proximas(latitude, longitude)
        
        if ctos_encontradas is None:
            await message.reply_text("❌ Ocorreu um erro ao acessar o banco de dados.")
        elif not ctos_encontradas:
            await message.reply_text("Nenhuma CTO foi encontrada no raio de 150 metros.")
        else:
            try:
                mapa_buffer = await criar_mapa_ctos(latitude, longitude, ctos_encontradas)
                if mapa_buffer:
                    linhas_ctos = [
                        f"- {escape(cto['cto'])} - <a href='https://maps.google.com/?q={cto['latitude']},{cto['longitude']}'>Rota</a>" 
                        for cto in ctos_encontradas
                    ]
                    nomes_ctos_com_link = "\n".join(linhas_ctos)
                    caption_text = f"✅ Encontrei {len(ctos_encontradas)} CTO(s) próximas:\n{nomes_ctos_com_link}"
                    
                    await context.bot.send_photo(
                        chat_id=message.chat.id,
                        photo=mapa_buffer,
                        caption=caption_text,
                        parse_mode=ParseMode.HTML
                    )
                else:
                    await message.reply_text("❌ Erro ao gerar o buffer do mapa.")
            except Exception as e:
                logger.error(f"Falha ao gerar o mapa para /ctos: {e}", exc_info=True)
                await message.reply_text("❌ Ocorreu um erro ao gerar o mapa.")
        return
        
    # Rota 2: Nova CTO (aceita link ou nativa)
    elif context.user_data.pop('waiting_for_location', False):
        await update.message.reply_text(f"📍 <b>Informações da Localização</b>\n\n"
                f"Latitude/Longitude: <code>{latitude}, {longitude}</code>\n"
                f"{escape(accuracy)}\n\n"
                f"<a href='https://maps.google.com/?q={latitude},{longitude}'>Abrir no Google Maps</a>\n\n""Enviando para o template...")
        # Recupera outras informações salvas.
        pop = context.user_data.pop('pop', None)
        olt_slot_pon = context.user_data.pop('olt_slot_pon', None)
        splitter = context.user_data.pop('splitter', None)

        if not all([pop, olt_slot_pon, splitter]):
            await update.message.reply_text("❌ Faltam informações para criar a CTO. Tente o comando /novaCTO novamente.")
            return
        
        webhook_link = await buscar_webhook_por_pop(pop)
        if not webhook_link:
            await update.message.reply_text(ErroP102)
            return

        # Prepara o payload para ser enviado ao sistema externo via webhook.
        olt, slot, pon = olt_slot_pon.split("/")
        payload = {"comando": "NovaCto", "olt": olt, "slot": slot, "pon": pon, "latitude": latitude, "longitude": longitude, "splitter": splitter, "id": update.effective_chat.id}
        data = await fetch_data(webhook_link, payload)
        await update.message.reply_text(data.get("mensagem", "Ocorreu um erro na resposta do servidor."))
        return
    
    # Rota 3 (Padrão): Apenas localização NATIVA
    else:
        if message and message.location:
            # Se for uma localização nativa, executa a Rota 3 normalmente.
            logger.info(f"Processando Rota 3: Localização nativa avulsa de {user.full_name}")
            accuracy = f"Precisão: {message.location.horizontal_accuracy:.0f} metros" if message.location.horizontal_accuracy else ""
            
            mensagem_final = (
                f"📍 <b>Informações da Localização</b>\n\n"
                f"Latitude/Longitude: <code>{latitude}, {longitude}</code>\n"
                f"{escape(accuracy)}\n\n"
                f"<a href='https://maps.google.com/?q={latitude},{longitude}'>Abrir no Google Maps</a>"
            )
            await message.reply_text(mensagem_final, parse_mode=ParseMode.HTML)
        

# --- Configuração de Logging para o Telegram ---

class TelegramHandler(logging.Handler):
    def emit(self, record):
        log_entry = self.format(record)
        try:
            loop = asyncio.get_running_loop()
            if loop.is_running():
                loop.create_task(send_log_to_telegram(log_entry))
        except RuntimeError:
            print(f"LOG_FALLBACK (no loop): {log_entry}")
            pass

telegram_handler = TelegramHandler()
telegram_handler.setLevel(logging.INFO)
telegram_formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
telegram_handler.setFormatter(telegram_formatter)
logger.addHandler(telegram_handler)

# Reduz o "ruído" de log de bibliotecas.
logging.getLogger("aiohttp").setLevel(logging.WARNING)
logging.getLogger("telegram").setLevel(logging.WARNING)
logging.getLogger("httpx").setLevel(logging.WARNING)
logging.getLogger("urllib3").setLevel(logging.WARNING)

# --- Funções Utilitárias para Manipulação de Arquivos ---

# NOVO: Função auxiliar para executar código bloqueante de forma assíncrona.
async def _run_blocking_io(func, *args, **kwargs):
    """Função auxiliar para executar código bloqueante em uma thread."""
    return await asyncio.to_thread(func, *args, **kwargs)

# Função para excluir arquivos com base em um padrão de nome.
async def ExcluirArquivos(caminho_arquivo):
    pasta, nome_base_ext = os.path.split(caminho_arquivo)
    nome_base, ext = os.path.splitext(nome_base_ext)
    arquivos_encontrados = glob.glob(os.path.join(pasta, f"{nome_base}*{ext}"))
    if arquivos_encontrados:
        for arquivo in arquivos_encontrados:
            try:
                await _run_blocking_io(os.remove, arquivo)
                print(f"✅ Arquivo excluído: {arquivo}")
            except Exception as e:
                print(f"❌ Erro ao excluir '{arquivo}': {e}")
    else:
        logger.info(f"Erro ao excluir um arquivo - ⚠️ Nenhuma versão do arquivo encontrada para exclusão: {caminho_arquivo}")


# Função para limpar o diretório raiz de arquivos com extensões específicas.
async def ExcluirArquivosporExtensao():
    diretorio = Path("")
    extensoes = [".xlsx", ".kml", ".kmz"]
    for arquivo in diretorio.iterdir():
        if arquivo.suffix in extensoes:
            await _run_blocking_io(arquivo.unlink)
            logger.info(f"Ajuste do diretório raiz - Excluído: {arquivo}")


# Função para converter um arquivo KML (XML para dados geográficos) em um arquivo XLSX (Excel).
def kml_to_xlsx(kml_file, xlsx_file):
    """Lógica síncrona para converter KML para XLSX."""
    tree = ET.parse(kml_file)
    root = tree.getroot()
    namespaces = {'kml': 'http://www.opengis.net/kml/2.2'}
    wb = Workbook()
    ws = wb.active
    ws.title = "Placemarks"
    ws.append(["PLACEMARK", "LATITUDE", "LONGITUDE"])
    for placemark in root.findall(".//kml:Placemark", namespaces):
        name_node = placemark.find("kml:name", namespaces)
        coord_node = placemark.find(".//kml:coordinates", namespaces)
        if name_node is not None and coord_node is not None and name_node.text and coord_node.text:
            coord_text = coord_node.text.strip()
            coord_parts = coord_text.split(",")
            if len(coord_parts) >= 2:
                lon, lat = coord_parts[0], coord_parts[1]
                ws.append([name_node.text.strip(), lat.strip(), lon.strip()])
    wb.save(xlsx_file)


# --- Funções de Busca em Arquivos JSON ---

# Função para listar cidades a partir de um arquivo JSON de configuração.
async def buscar_webhook_por_pop(pop: str) -> str | None:
    """Busca o link do webhook associado a um POP de forma segura."""
    try:
        async with aiofiles.open(FILENAME_WEBHOOK, 'r', encoding='utf-8') as f:
            dados = json.loads(await f.read())
        for entry in dados:
            if entry.get("POP", "").upper() == pop.upper():
                return entry.get("WEBHOOK_LINK")
        return None
    except FileNotFoundError:
        logger.error(f"Erro CRÍTICO: O ficheiro de configuração '{FILENAME_WEBHOOK}' não foi encontrado.")
        return None
    except json.JSONDecodeError:
        logger.error(f"Erro CRÍTICO: O ficheiro '{FILENAME_WEBHOOK}' contém um JSON inválido.")
        return None
    except Exception as e:
        logger.error(f"Erro inesperado ao ler o ficheiro de webhook: {e}")
        return None

async def buscar_cidade_por_pop(pop: str) -> str | None:
    """Busca o nome da cidade associada a um POP de forma segura."""
    try:
        async with aiofiles.open(FILENAME_WEBHOOK, 'r', encoding='utf-8') as f:
            dados = json.loads(await f.read())
        for entry in dados:
            if entry.get("POP", "").upper() == pop.upper():
                return entry.get("CIDADE")
        return None
    except FileNotFoundError:
        logger.error(f"Erro CRÍTICO: O ficheiro de configuração '{FILENAME_WEBHOOK}' não foi encontrado.")
        return None
    except json.JSONDecodeError:
        logger.error(f"Erro CRÍTICO: O ficheiro '{FILENAME_WEBHOOK}' contém um JSON inválido.")
        return None
    except Exception as e:
        logger.error(f"Erro inesperado ao buscar cidade por POP: {e}")
        return None


async def ListaCidades():
    """Lista cidades a partir de um ficheiro JSON de forma segura."""
    try:
        async with aiofiles.open(FILENAME_WEBHOOK, 'r', encoding='utf-8') as f:
            dados = json.loads(await f.read())
        return "\n".join([f"{i+1}. {c.get('POP', '')} - {c.get('CIDADE', '')}" for i, c in enumerate(dados)])
    except FileNotFoundError:
        logger.error(f"Erro CRÍTICO: O ficheiro de cidades '{FILENAME_WEBHOOK}' não foi encontrado.")
        return "Arquivo de cidades não encontrado."
    except json.JSONDecodeError:
        logger.error(f"Erro CRÍTICO: O ficheiro de cidades '{FILENAME_WEBHOOK}' contém um JSON inválido.")
        return "Arquivo de cidades corrompido."
    except Exception as e:
        logger.error(f"Erro inesperado ao ler o ficheiro de cidades: {e}")
        return "Erro ao ler o arquivo de cidades."



# Busca a configuração de um diretório em um arquivo JSON específico.
async def buscar_dir_drive():
    try:
        async with aiofiles.open("config_drive.json", "r", encoding="utf-8") as f:
            dados = json.loads(await f.read())
            return dados.get("diretorio", "❌ Diretório não encontrado no arquivo.")
    except FileNotFoundError:
        return "❌ Arquivo de configuração não encontrado."

# --- Funções de Manipulação de Arquivos para o Drive ---

# Função para mover um arquivo para um diretório, com versionamento automático.
def EnviaArquivosDrive(dirarquivo, xlsx_file):
    if not os.path.exists(dirarquivo):
        os.makedirs(dirarquivo)

    nome_base, ext = os.path.splitext(os.path.basename(xlsx_file))
    caminho_destino = os.path.join(dirarquivo, os.path.basename(xlsx_file))
    contador = 1
    while os.path.exists(caminho_destino):
        contador += 1
        novo_nome = f"{nome_base}_v{contador}{ext}"
        caminho_destino = os.path.join(dirarquivo, novo_nome)
    
    if os.path.exists(xlsx_file):
        shutil.move(xlsx_file, caminho_destino)


# Encontra o primeiro arquivo .kml ou .kmz em um dado diretório.
def encontrar_arquivo_kml_kmz(DirArquivo):
    if not os.path.exists(DirArquivo):
        logger.info(f"Encontrar arquivo - ❌ Diretório não encontrado: {DirArquivo}")
        return None

    # Itera sobre os arquivos no diretório.
    for arquivo in os.listdir(DirArquivo):
        if arquivo.endswith((".kml", ".kmz")):
            return os.path.join(DirArquivo, arquivo) 
    return None # Retorna None se nenhum arquivo for encontrado.

# Função para extrair o arquivo .kml de dentro de um arquivo .kmz.
def extract_kml_from_kmz(kmz_file, extract_to):
    with zipfile.ZipFile(kmz_file, 'r') as kmz:
        for file in kmz.namelist():
            if file.endswith('.kml'):
                kmz.extract(file, extract_to)
                kml_file = os.path.join(extract_to, file)
                new_kml_file = os.path.join(extract_to, os.path.splitext(os.path.basename(kmz_file))[0] + '.kml')
                os.rename(kml_file, new_kml_file)
                return new_kml_file
    return None


# Converte uma planilha (XLSX) para um arquivo KML.
def converter_planilha(CaminhoXLSX, CaminhoKML, NomePlanilha, IconeUrl):
    workbook = openpyxl.load_workbook(CaminhoXLSX)
    if NomePlanilha not in workbook.sheetnames:
        raise FileNotFoundError(f"Planilha '{NomePlanilha}' não encontrada em {CaminhoXLSX}")
    
    sheet = workbook[NomePlanilha]
    kml = simplekml.Kml()

    for row in sheet.iter_rows(min_row=3, values_only=True):
        nome, lat, lon = row[0], row[1], row[2]
        if nome and lat and lon:
            pnt = kml.newpoint(name=str(nome), coords=[(lon, lat)])
            pnt.style.iconstyle.icon.href = IconeUrl
            pnt.style.iconstyle.scale = 1.5
    
    kml.save(CaminhoKML)
    workbook.close()


# Função para copiar dados de uma planilha de origem para uma planilha de destino (template).
def DE_KMZ_BASE_PARA_TEMPLATE(arquivo_origem, arquivo_destino):
    try:
        wb_origem = openpyxl.load_workbook(arquivo_origem)
        sheet_origem = wb_origem.active
        wb_destino = openpyxl.load_workbook(arquivo_destino)
        if "KMZ" not in wb_destino.sheetnames:
            sheet_destino = wb_destino.create_sheet("KMZ")
        else:
            sheet_destino = wb_destino["KMZ"]
        for row_idx, row in enumerate(sheet_origem.iter_rows(min_row=3, values_only=True), start=3):
            sheet_destino[f"A{row_idx}"] = row[0]
            sheet_destino[f"B{row_idx}"] = row[1]
            sheet_destino[f"C{row_idx}"] = row[2]
        wb_destino.save(arquivo_destino)
        print("✅ Dados copiados com sucesso para a planilha 'KMZ'!")

    except Exception as e:
        logger.error(f"❌ Erro ao copiar os dados: {e}")


# Verifica a existência de um arquivo de template baseado no POP e em uma convenção de nome.
async def VerificarTemplatemporPOP(DirTemplate, PopInformado_user, update):
    if os.path.exists(DirTemplate):
        for arquivo in os.listdir(DirTemplate):
            if arquivo.startswith("TEMPLATE REDES") and arquivo.endswith(".gsheet"):
                partes = arquivo.replace(".gsheet", "").split()
                if len(partes) >= 3:
                    pop_do_arquivo = partes[2]
                    
                    if pop_do_arquivo == PopInformado_user:
                        caminho_arquivo = os.path.join(DirTemplate, arquivo)
                        logger.info(f"handle_mensagem - POP: {pop_do_arquivo} - Arquivo encontrado: {caminho_arquivo}")
                    else:
                        logger.info(f"handle_mensagem - POP: {pop_do_arquivo} não corresponde ao esperado.")
    else:
        await update.message.reply_text(
            f"Erro ao acessar o template no diretório do drive.\n\n"
            f"| Informações recebidas:\nCaminho recebido do Template:\n{DirTemplate}"
        )


# Handler para o comando /ajuda.
async def ajuda(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """
    Envia a mensagem de ajuda completa, seguindo o padrão de formatação de texto original.
    """
    message = update.message or update.edited_message
    if not message:
        return

    user = update.effective_user
    logger.info(f"Comando /ajuda solicitado por: {user.full_name} ({user.id})")

    # Lista de comandos formatada no estilo original
    comandos = [
        "| Ajuda - BOT-ALEXO",

        "\n\n- Atividades 🌟",
        "    /atividades <POP>",
        "    Verifica atividades e gera de acordo com as atividades pendentes no template.",
        "    EX: /atividades POP",

        "\n\n- Checar 🔍",
        "    /checar <CTO> <FSAN>",
        "    Verifica OLT/SLOT/PON de um cliente na CTO.",
        "    EX: /checar CTO-001 FHTT0000000",

        "\n\n- Localizar CTO 📍",
        "    /localizar <CTO>",
        "    Retorna a localização geográfica de uma CTO.",
        "    EX: /localizar CTO-001",
        
        "\n\n- CTOs Próximas 🗺️",
        "    /ctos",
        "    Retorna a localização para de CTOs próximas.",
        "    EX: /ctos",

        "\n\n- Listar IDs 📋",
        "    /listarIDs <POP> <OLT/SLOT/PON>",
        "    Lista IDs de CTOs disponíveis em uma PON.",
        "    EX: /listarIDs POP 1/1/1",
        
        "\n\n- Id",
        "    /id"
        "    Informa o ID do seu usuário e do chat atual.",

        "\n\n- Input 📝",
        "    /input <CTO> <SPLITTER>",
        "    Inputa data e splitter no template.",
        "    EX: /input CTO-001 1/16",

        "\n\n- Informações",
        "    Exibe informações, versão e criadores do bot."
        "    EX: /info",

        "\n\n- Insert ➡️",
        "    /insert <CTO> <OLT/SLOT/PON>",
        "    Inputa OLT/SLOT/PON na aba 'checar' do template.",
        "    EX: /insert CTO-001 1/1/1",

        "\n\n- Nova CTO ➕",
        "    /novaCTO <POP> <OLT/SLOT/PON> <SPLITTER>",
        "    Adicionar uma CTO que não está no KMZ.",
        "    EX: /novaCTO CTO 1/1/1 1/16",
        
        "\n\n- Versão",
        "    /versao",
        "    Apresenta a versão atual e os créditos.",

        "\n\n- Ajuda Administração",
        "    /ajudaadm",
        "    Lista os Comandos de Administrador"
    ]

    comandos_texto = "\n".join(comandos)
    logger.info(f"/Ajuda - Usuário:{user.first_name} {user.last_name}, Grupo:{update.effective_chat.title}")
    await context.bot.send_message(chat_id=update.effective_chat.id, text=comandos_texto)

# --- Função de Comunicação com Webhook ---

# Função para enviar dados para um webhook e receber a resposta.
async def fetch_data(webhook_link, payload):
    try:
        async with aiohttp.ClientSession() as session:
            async with session.post(webhook_link, json=payload) as response:
                response_data = await response.json()

                if response.status == 200:
                    logger.info(f"Google App Script - Resposta: {response_data}")
                    return response_data
                else:
                    logger.error(f"Erro ao conectar ao Apps Script: {response.status} - {response.reason}")
                    return {
                        "status": "error",
                        "message": f"Erro ao conectar ao servidor: {response.reason}."
                    }

    except aiohttp.ClientError as client_error:
        logger.error(f"Erro de cliente HTTP: {client_error}")
        return {"status": "error", "message": "Erro de comunicação com o servidor."}

    except Exception as e:
        logger.error(f"/fetch_data - Exceção inesperada: {e}")
        return {"status": "error", "message": f"Erro inesperado: {str(e)}"}


# --- Handlers de Comando ---

# Comando /atividades.
@check_permission
async def atividades(update: Update, context: ContextTypes.DEFAULT_TYPE):
    # Valida se o argumento <POP> foi informado.
    if len(context.args) < 1:
        await update.message.reply_text(text=ErroP101)
        return
    pop = context.args[0].upper().split('-')[0]
    
    # Busca o link do webhook correspondente ao POP.
    webhook_link = await buscar_webhook_por_pop(pop)
    if webhook_link is None:
        await update.message.reply_text(ErroP102)
        return

    # Monta o payload para a requisição.
    payload = {"comando": "Atividades", "id": update.effective_chat.id}
    logger.info(f"RECEBIDO: /Atividades - POP:{pop} - Usuário:{update.effective_user.first_name}")

    try:
        # Envia a requisição para o webhook.
        data = await fetch_data(webhook_link, payload)
    except Exception as e:
        logger.error(f"Erro ao buscar dados: {e}")
        await update.message.reply_text(text="⚠️ Erro ao processar a solicitação.")
        return
    
    # Processa a resposta do webhook.
    if data.get("status") == "sucesso":
        await context.bot.send_message(chat_id=update.effective_chat.id, text=f"{data.get('mensagem')}")
        logger.info(f"Atividade: {data.get('mensagem')}")
    else:
        # Caso o status não seja "sucesso", exibe uma mensagem de erro detalhada.
        ErroWH104 = (
            "WH104.\n\n| VERIFICAR SE A SIGLA DO POP FOI INFORMADO CORRETAMENTE!"
            "\n\nCaso persistir, informar o erro à equipe interna com urgência!"
            "\n\nCONTATOS:\n    - @J_Ayrton\n    - @AlexandreBarros_Desktop"
        )
        error_message = data.get("mensagem", ErroWH104)
        logger.error(f"ERRO WH104: COMANDO /Atividades - POP:{pop} - Usuário:{update.effective_user.first_name}")
        await context.bot.send_message(chat_id=update.effective_chat.id, text=f"⚠️ Erro 1: {error_message}")

    return webhook_link


# Comando /checar.
@check_permission
async def checar(update: Update, context: ContextTypes.DEFAULT_TYPE):
    # Valida a quantidade de argumentos recebidos.
    if len(context.args) < 2:
        await update.message.reply_text(text=ErroF101 if len(context.args) < 2 else ErroP101)
        return

    cto, fsan = context.args[:2]
    cto = str(cto.upper())
    pop = cto.split('-')[0]

    # Validações de formato para os argumentos.
    if cto.count('-') != 1:
        await update.message.reply_text(text=ErroC101)
        return
    if '/' in fsan or '-' in fsan:
        await update.message.reply_text(text=ErroF102)
        return

    # Busca o webhook e, se não encontrar, envia erro.
    webhook_link = await buscar_webhook_por_pop(pop)
    if webhook_link is None:
        await update.message.reply_text(ErroP102)
        return

    payload = {"comando": "Checar", "cto": cto, "fsan": fsan}
    logger.info(f"/Checar recebido - CTO: {cto}, FSAN: {fsan} - Usuário:{update.effective_user.first_name}")
    
    # Envia os dados e processa a resposta.
    data = await fetch_data(webhook_link, payload)
    if data.get("status") == "sucesso":
        await update.message.reply_text(text=f"{data.get('confirmacao')}")
    else:
        await update.message.reply_text(text=f" 6: {data.get('mensagem')}")
    return webhook_link


# Comando /localizar.
@check_permission
async def localizar_cto(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if len(context.args) < 1:
        await update.message.reply_text(text=ErroP101)
        return

    cto = context.args[0].upper()

    if cto.count('-') != 1:
        await update.message.reply_text(text=ErroC101)
        return

    pop = cto.split('-')[0]
    webhook_link = await buscar_webhook_por_pop(pop)
    if webhook_link is None:
        await update.message.reply_text(ErroP102)
        return

    payload = {"comando": "Localizar", "cto": cto}
    logger.info(f"/Localizar recebido - POP: {pop}, CTO: {cto} - Usuário:{update.effective_user.first_name}")
    data = await fetch_data(webhook_link, payload)

    if data.get("status") == "sucesso":
        await update.message.reply_text(text=f"{data.get('mensagem')}")
    else:
        await update.message.reply_text(text="⚠️ CTO NÃO ENCONTRADO!")


# Comando /Exibircidade.
@check_permission
async def exibircidade(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Exibe a lista de cidades e POPs configurados."""
    message = update.message or update.edited_message
    if not message:
        return

    # Await é necessário pois ListaCidades é uma função async
    cidades = await ListaCidades()
    
    await message.reply_text(text=f"🌆 Cidades disponíveis:\n\n{cidades}")
    logger.info(f"/ExibirCidade recebido - Usuário: {update.effective_user.full_name}")
    

# Comando /input
@check_permission
async def input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if len(context.args) < 2:
        await update.message.reply_text(text=ErroS101 if len(context.args) > 0 else ErroP101)
        return

    cto, splitter = context.args[:2]
    
    # Validações de formato para CTO e splitter.
    if cto.count('-') != 1:
        await update.message.reply_text(text=ErroC101)
        return
    if splitter.count('/') != 1:
        await update.message.reply_text(text=ErroN102)
        return

    cto = cto.upper()
    pop = cto.split('-')[0]
    webhook_link = await buscar_webhook_por_pop(pop)
    if webhook_link is None:
        await update.message.reply_text(ErroP102)
        return

    # Validação específica para o valor do splitter.
    splitters_validos = {"16", "8", "4"}
    splitter_final = splitter.split("/")[-1]
    if splitter_final not in splitters_validos:
        await update.message.reply_text(text="❌ SPLITTER inválido! Use apenas 1/16, 1/8, 1/4.")
        return

    payload = {"comando": "Input", "cto": cto, "splitter": splitter_final}
    logger.info(f"/Input recebido - POP: {pop}, CTO: {cto} - Usuário:{update.effective_user.first_name}")
    
    try:
        data = await fetch_data(webhook_link, payload)
        if data.get("status") == "sucesso":
            await update.message.reply_text(text=f"{data.get('confirmacao')}")
        else:
            await update.message.reply_text(text=f" ⚠️ Erro 5: {data.get('mensagem')}")
    except Exception as e:
        logger.error(f"/Input recebido - POP: {pop}, CTO: {cto}, {e} - Usuário:{update.effective_user.first_name}")
        await update.message.reply_text(text="⚠️ Erro interno ao processar sua solicitação. Tente novamente mais tarde.")
        return    

# --- Comandos Administrativos e de Informação ---

# Comando de ajuda administrativa.
@check_permission
async def ajudaadm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    chat_id = update.effective_chat.id
    chat_title = update.effective_chat.title or "Chat Privado"
    
    # Monta uma única string de texto com todos os comandos administrativos e suas descrições.
    comandos =[
        "| AjudaAdm:"
        "\n\n>>> Principais comandos"
        "\n\n- Cadastrar",
        "    Cadastrar <CARGO>",
        "    (Admin) Gera um link de convite para um novo usuário.",
        "    EX: /cadastrar Tecnico",
        
        "\n\n- Exibir Cidades",
        "    /exibircidade",
        "    Lista todas as cidades e POPs configurados.",

        "\n\n- Adicionar Template"
        "    /AdicionarTemplate <CIDADE> <POP> <WEBHOOK>",
        "    Adiciona um novo link de template cidade.",
        "    EX: /AdicionarTemplate RIO_CLARO POP HTTP://...",

        "\n\n- Excluir Template"
        "    /ExcluirTemplate <POP>",
        "    Remove uma configuração de template pelo POP.",
        "    /ExcluirTemplate POP",

        "\n\n- Configuração Drive",
        "    /configdrive <CAMINHO>",
        "    Define o diretório raiz do Drive local.",
        "    /configdrive G:/MEU DRIVE/FASTERNET...",

        "\n\n- Compartilhar Webhook",
        "    /CWH",
        "    Envia o arquivo de configuração WebHook.json.",

        "\n\n- Converter Arquivos",
        "    /convert",
        "    Converter um arquivo KML/KMZ em XLSX",
        
        "\n\n- Baixa Arquivos KMZ",
        "    /baixarkmz <POP>",
        "    Baixa o arquivo KMZ/KML do Drive.",
        "    /baixarkmz POP",

        "\n\n- Gerar KMZ",
        "    /gerarkmzatualizado <POP>",
        "    Gera um arquivo KML base a partir do template.",
        "    /gerarkmzatualizado POP",
        
        "\n\n- Listar Admins",
        "    /listar_admins",
        "    Exibe a lista de administradores do grupo."
    ]
    
    comandos_texto = "\n".join(comandos)
    await context.bot.send_message(chat_id=chat_id, text=comandos_texto)
    logger.info(f"/ajudaadm - Usuário:{user.first_name}, Grupo:{chat_title}")
    
# Comando /CWH (Compartilhar WebHook).
@check_permission
async def CWH(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    chat_id = update.effective_chat.id
    chat_title = update.effective_chat.title or "Chat Privado"
    logger.info(f"/CWH - Usuário:{user.first_name}, Grupo:{chat_title}")
    try:
        # Abre o arquivo de forma assíncrona para leitura em bytes ('rb')
        async with aiofiles.open('WebHook.json', 'rb') as f:
            await context.bot.send_document(chat_id=chat_id, document=f)
    except FileNotFoundError:
        await update.message.reply_text("❌ Arquivo 'WebHook.json' não encontrado.")
    except Exception as e:
        logger.error(f"Erro ao enviar WebHook.json: {e}")
        await update.message.reply_text("❌ Ocorreu um erro ao enviar o arquivo.")

    
# Comando para Adicionar um novo template ao arquivo de configuração WebHook.json.
@check_permission
async def adicionartemplate(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    if len(context.args) < 3:
        await update.message.reply_text(
            text=("❌ Formato inválido!\n\n"
                  "Use: /AdicionarTemplate <CIDADE> <POP> <WEBHOOK>\n\n"
                  "Exemplo:\n/AdicionarTemplate Rio_Claro RCA https://script.google.com/macros...")
        )
        return

    CIDADE_ID, POP, WEBHOOK_LINK = context.args[:3]
    CIDADE_ID = CIDADE_ID.upper().replace("-", "_") # Padroniza o nome da cidade
    POP = POP.upper()
    
    novo_dado = {'CIDADE': CIDADE_ID, 'POP': POP, 'WEBHOOK_LINK': WEBHOOK_LINK}
    
    try:
        dados_existentes = []
        if await _run_blocking_io(os.path.exists, FILENAME_WEBHOOK):
            async with aiofiles.open(FILENAME_WEBHOOK, 'r', encoding='utf-8') as f:
                try:
                    dados_existentes = json.loads(await f.read())
                except json.JSONDecodeError:
                    pass # Arquivo existe mas está vazio ou corrompido

        dados_existentes.append(novo_dado)
        
        async with aiofiles.open(FILENAME_WEBHOOK, 'w', encoding='utf-8') as f:
            await f.write(json.dumps(dados_existentes, ensure_ascii=False, indent=4))

        cidades = await ListaCidades() # Await na chamada da função async
        await update.message.reply_text(text=f"✅ Novo template adicionado:\n\n- CIDADE: {CIDADE_ID}\n- POP: {POP}\n- WEBHOOK: {WEBHOOK_LINK}")
        await update.message.reply_text(text=f"Lista de cidades existentes:\n\n{cidades}")
        logger.info(f"/AdicionarTemplate - CIDADE:{CIDADE_ID}, POP:{POP} - Usuário:{user.first_name}")

    except Exception as e:
        logger.error(f"Erro em /AdicionarTemplate: {e}")
        await update.message.reply_text("❌ Ocorreu um erro ao adicionar o template.")

# Comando para excluir um template do arquivo de configuração.
@check_permission
async def excluirtemplate(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if len(context.args) < 1:
        await update.message.reply_text(text="❌ Formato inválido!\n\nUse: /ExcluirTemplate <POP>")
        return

    pop_a_excluir = context.args[0].upper()
    logger.info(f"/ExcluirTemplate - POP para excluir: {pop_a_excluir} - Usuário: {update.effective_user.first_name}")

    try:
        if not await _run_blocking_io(os.path.exists, FILENAME_WEBHOOK):
            await update.message.reply_text("❌ Arquivo de configuração 'WebHook.json' não encontrado.")
            return

        async with aiofiles.open(FILENAME_WEBHOOK, 'r', encoding='utf-8') as f:
            content = await f.read()
            dados = json.loads(content) if content else []

        dados_atualizados = [item for item in dados if item.get('POP') != pop_a_excluir]

        # Verifica se algum item foi removido comparando o tamanho das listas.
        if len(dados) == len(dados_atualizados):
            await update.message.reply_text(text=f"⚠️ O POP '{pop_a_excluir}' não foi encontrado na lista.")
        else:
            async with aiofiles.open(FILENAME_WEBHOOK, 'w', encoding='utf-8') as f:
                await f.write(json.dumps(dados_atualizados, indent=4, ensure_ascii=False))
            
            await update.message.reply_text(text=f"✅ O POP '{pop_a_excluir}' foi excluído com sucesso!")
            cidades = await ListaCidades() # Await na chamada da função async
            await update.message.reply_text(text=f"Lista de cidades existentes:\n\n{cidades}")
            
    except (FileNotFoundError, json.JSONDecodeError):
        await update.message.reply_text("❌ Arquivo de configuração 'WebHook.json' não encontrado ou corrompido.")
    except Exception as e:
        logger.error(f"Erro inesperado ao excluir template: {e}", exc_info=True)
        await update.message.reply_text("❌ Ocorreu um erro interno ao tentar excluir o template.")


# Comando público /id.
async def id(update: Update, context: ContextTypes.DEFAULT_TYPE):
    # Obtém informações de IDs do chat e do usuário.
    chat_id = update.effective_chat.id
    chat_title = update.effective_chat.title
    user_id = update.effective_user.id
    
    logger.info(f"/id - Usuário:{update.effective_user.first_name}, Grupo:{chat_title}")
    # Envia uma mensagem com as informações coletadas.
    await update.message.reply_text(f"\nA ID deste grupo é: {chat_id}, "
                                      f"\nNome do grupo: {chat_title},"
                                      f"\nID do Usuario: {user_id}")

# Comando público /info.
async def info(update: Update, context: ContextTypes.DEFAULT_TYPE):
    Inf = (
        "| Nome do BOT: Alexo"
        "\n\n - Alexo tem o intuito de ser um auxílio para os usuários técnicos, back-offices e internos, com a capacidade de gerar de editar plalhas inopputando informações direto do chat, assim reduzindo as margens se erros na inputação de diversos procedimentos por todas as equipes."
        f"\n\nVersão: {__version__}"
        f"\n\nCriador: {__author__}"
        f"\nCréditos: {__credits__}"
    )
    logger.info(f"/Info - Usuário:{update.effective_user.first_name}, Grupo:{update.effective_chat.title}")
    await update.message.reply_text(Inf)

# --- Handlers de Comando Adicionais ---

# Comando /listarIDs.
@check_permission
async def listarids(update: Update, context: ContextTypes.DEFAULT_TYPE):
    # Valida se os argumentos <POP> e <OLT/SLOT/PON> foram fornecidos.
    if len(context.args) < 2:
        await update.message.reply_text(text=ErroS101 if len(context.args) > 0 else ErroP101)
        return
    
    pop, OLT_SLOT_PON = context.args[:2]
    pop = pop.upper().split('-')[0]
    if OLT_SLOT_PON.count('/') != 2:
        await update.message.reply_text(text=ErroN102)
        return

    # Extrai os componentes olt, slot e pon.
    olt, slot, pon = OLT_SLOT_PON.split("/")
    payload = {"comando": "ListarIds", "olt": olt, "slot": slot, "pon": pon}
    logger.info(f"/ListarIDs - OLT:{olt}, SLOT:{slot}, PON:{pon} - Usuário:{update.effective_user.first_name}")

    webhook_link = await buscar_webhook_por_pop(pop) # CORRIGIDO: Adicionado await
    if webhook_link is None:
        await update.message.reply_text(ErroP102)
        return
    
    data = await fetch_data(webhook_link, payload)

    if data.get("status") == "sucesso":
        ctos = data.get('mensagem', []) # Garante que 'ctos' seja uma lista, mesmo se a chave não existir.
        ctos_com_contador = [f"{i+1}. {cto}" for i, cto in enumerate(ctos)]
        ctos_com_contador_str = '\n'.join(ctos_com_contador)
        await update.message.reply_text(text=f"IDs disponiveis:\n\n{ctos_com_contador_str}\n\n| Sempre use o Id da CTO de número [1]")
    else:
        await update.message.reply_text(text=f"⚠️ Erro 4: {data.get('mensagem')}")

# Comando /insert.
@check_permission
async def insert(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if len(context.args) < 2:
        await update.message.reply_text(text=ErroN101 if len(context.args) > 0 else ErroP101)
        return
    
    CTO, OLT_SLOT_PON = context.args[:2]
    CTO = str(CTO.upper())
    POP = CTO.split('-')[0]

    if CTO.count('-') != 1:
        await update.message.reply_text(text=ErroC101)
        return
    if OLT_SLOT_PON.count('/') != 2:
        await update.message.reply_text(text=ErroN102)
        return
    
    olt, slot, pon = OLT_SLOT_PON.split("/")
    payload = {"comando": "Insert", "cto": CTO, "olt": olt, "slot": slot, "pon": pon}
    logger.info(f"/Insert - CTO:{CTO}, PON:{OLT_SLOT_PON} - Usuário:{update.effective_user.first_name}")

    webhook_link = await buscar_webhook_por_pop(POP)
    
    if webhook_link is None:
        await update.message.reply_text(ErroP102)
        return
    
    data = await fetch_data(webhook_link, payload)
    
    if data.get("status") == "sucesso":
        await update.message.reply_text(text=f"{data.get('mensagem')}")
    else:
        await update.message.reply_text(text=f"⚠️ Erro 3: {data.get('mensagem')}")

# Comando /novaCTO.
@check_permission
async def novacto(update: Update, context: ContextTypes.DEFAULT_TYPE):
    # Validação robusta para garantir que todos os 3 argumentos foram fornecidos.
    if len(context.args) < 3:
        await update.message.reply_text(text=ErroS101 if len(context.args) > 1 else ErroN101 if len(context.args) > 0 else ErroP101)
        return
    
    pop, olt_slot_pon, splitter = context.args
    pop = pop.split('-')[0]

    if olt_slot_pon.count('/') != 2 or splitter.count('/') != 1:
        await update.message.reply_text(text=ErroN102)
        return

    # Validação específica para o valor do splitter.
    splitters_validos = {"16", "8", "4"}
    splitter_final = splitter.split("/")[-1]
    if splitter_final not in splitters_validos:
        await update.message.reply_text(text="❌ SPLITTER inválido! Use apenas 1/16, 1/8, 1/4.")
        return
        
    await update.message.reply_text(text="📍 Por favor, envie a localização da CTO que deseja adicionar.")

    # Armazena os dados já coletados e define uma flag de estado em 'user_data'.
    context.user_data['waiting_for_location'] = True
    context.user_data['pop'] = pop
    context.user_data['olt_slot_pon'] = olt_slot_pon
    context.user_data['splitter'] = splitter_final
    logger.info(f"/NovaCTO - POP:{pop}, PON:{olt_slot_pon}, SPL:{splitter} - Usuário:{update.effective_user.full_name}")

# Comando /convert.
@check_permission
async def convert(update: Update, context: ContextTypes.DEFAULT_TYPE):
    # Define a flag de estado para indicar que o bot está aguardando um arquivo.
    context.user_data['waiting_for_file'] = True
    await update.message.reply_text("Por favor, envie o arquivo KML/KMZ que você deseja converter.")


# --- Handler de Arquivo ---

# Handler para processar documentos recebidos.
async def handle_arquivo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not context.user_data.get('waiting_for_file'):
        return

    if not (update.message and update.message.document):
        await update.message.reply_text("❌ Não foi possível identificar o arquivo. Por favor, envie um arquivo válido.")
        return

    context.user_data.pop('waiting_for_file', None)
    document = update.message.document
    file_name = document.file_name
    
    await update.message.reply_text(f"📥 Recebendo arquivo '{file_name}'... Por favor, aguarde.")

    try:
        file = await context.bot.get_file(document.file_id)
        # file.download_to_drive é síncrono, precisa ser executado em uma thread.
        await _run_blocking_io(file.download_to_drive, f"{file_name}")
        logger.info(f"Arquivo Recebido - Arquivo:{file_name} - Usuário:{update.effective_user.first_name}")

        xlsx_file = None
        if file_name.endswith('.kml'):
            xlsx_file = file_name.replace('.kml', '.xlsx')
            await _run_blocking_io(kml_to_xlsx, file_name, xlsx_file)
        elif file_name.endswith('.kmz'):
            kml_file = await _run_blocking_io(extract_kml_from_kmz, file_name, "")
            if kml_file:
                xlsx_file = kml_file.replace('.kml', '.xlsx')
                await _run_blocking_io(kml_to_xlsx, kml_file, xlsx_file)
            else:
                await update.message.reply_text("❌ Não foi possível extrair o arquivo KML do KMZ.")
                return
        else:
            await update.message.reply_text(f"❌ O arquivo '{file_name}' não é um KML ou KMZ válido.")
            return

        if xlsx_file and await _run_blocking_io(os.path.exists, xlsx_file):
            await update.message.reply_text(f"✅ Arquivo convertido para '{xlsx_file}' com sucesso!")
            async with aiofiles.open(xlsx_file, 'rb') as f:
                await context.bot.send_document(chat_id=update.effective_chat.id, document=f)
            
            await update.message.reply_text("Digite uma opção:\n\n[0] - Sair\n\n[1] - Salvar no drive\n[2] - Salvar no drive e inputar no template")
            context.user_data['MsgUser_ApplyPointTemplates'] = True
            context.user_data['xlsx_file'] = xlsx_file
    except Exception as e:
        logger.error(f"Erro em handle_arquivo: {e}", exc_info=True)
        await update.message.reply_text("❌ Ocorreu um erro ao processar o arquivo.")

# --- Comandos de Configuração e Manipulação de Arquivos do Drive ---

# Comando /configdrive.
@check_permission
async def configdrive(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not context.args:
        await update.message.reply_text("❌ Você precisa informar um diretório! Exemplo: /configdrive nome_da_pasta")
        return

    DirDrive = context.args[0]
    dados = {"diretorio": DirDrive}

    try:
        async with aiofiles.open("config_drive.json", "w", encoding="utf-8") as f:
            await f.write(json.dumps(dados, ensure_ascii=False, indent=4))
        await update.message.reply_text(f"✅ Diretório salvo: {DirDrive}")
    except Exception as e:
        logger.error(f"Erro ao salvar config_drive.json: {e}")
        await update.message.reply_text("❌ Ocorreu um erro ao salvar a configuração.")


# Comando /baixarkmz 
async def baixarkmz(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not context.args:
        await update.message.reply_text("❌ Você precisa informar um POP válido!")
        return
        
    pop = context.args[0].upper().split('-')[0]
    
    try:
        NomeCidade = await buscar_cidade_por_pop(pop)
        if not NomeCidade:
            await update.message.reply_text(f"❌ Cidade para o POP '{pop}' não encontrada.")
            return

        LinkDrive = await buscar_dir_drive()
        if "❌" in LinkDrive:
            await update.message.reply_text(LinkDrive)
            return
            
        Pastakmz = f"{NomeCidade.replace('-', ' ')}/kmz e kml"
        dirarquivo = os.path.join(LinkDrive, Pastakmz)

        arquivo_path = await _run_blocking_io(encontrar_arquivo_kml_kmz, dirarquivo)
        
        if arquivo_path:
            await update.message.reply_text(f"Enviando arquivo: {os.path.basename(arquivo_path)}")
            async with aiofiles.open(arquivo_path, "rb") as f:
                await update.message.reply_document(document=f)
        else:
            await update.message.reply_text(f"❌ Nenhum arquivo KML/KMZ encontrado no diretório: {dirarquivo}")
    except Exception as e:
        logger.error(f"Erro no comando /baixarkmz: {e}", exc_info=True)
        await update.message.reply_text("Ocorreu um erro ao buscar o arquivo.")


# Handler para gerar um KML "base" a partir de dados de uma planilha de template.
async def gerarkmzatualizado(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    if not context.args:
        await update.message.reply_text("❌ Você precisa informar um POP válido!")
        return

    POP_ConvertKML = context.args[0].upper().split('-')[0]
    
    try:
        NomeCidade_ConvertKML = await buscar_cidade_por_pop(POP_ConvertKML)
        if not NomeCidade_ConvertKML:
            await update.message.reply_text(f"❌ Não foi possível encontrar a cidade para o POP: {POP_ConvertKML}")
            return
        
        LinkDrive = await buscar_dir_drive()
        if "❌" in LinkDrive:
            await update.message.reply_text(LinkDrive)
            return

        # ATENÇÃO: O caminho para o template pode precisar de ajuste. Assumindo .xlsx
        CaminhoXLSX = os.path.join(LinkDrive, NomeCidade_ConvertKML.replace("-", " "), "CEP CTO", f"TEMPLATE REDES {POP_ConvertKML}.xlsx") 
        NomePlanilha = "KMZ"
        IconeUrl = "http://maps.google.com/mapfiles/kml/shapes/placemark_circle.png"
        CaminhoKML = f"{POP_ConvertKML} - {NomeCidade_ConvertKML} - KMZ BASE.kml"
        
        if not await _run_blocking_io(os.path.exists, CaminhoXLSX):
             await update.message.reply_text(f"❌ Arquivo de template não encontrado em: {CaminhoXLSX}")
             return

        await context.bot.send_message(chat_id, "Gerando KML, aguarde...")
        await _run_blocking_io(converter_planilha, CaminhoXLSX, CaminhoKML, NomePlanilha, IconeUrl)
        
        async with aiofiles.open(CaminhoKML, "rb") as f:
            await context.bot.send_document(
                chat_id=chat_id,
                document=f,
                caption=f"Arquivo KML gerado com sucesso:\n   {os.path.basename(CaminhoKML)}"
            )
    except Exception as e:
        logger.error(f"Erro em /gerarkmzatualizado: {e}", exc_info=True)
        await update.message.reply_text("Ocorreu um erro ao gerar o arquivo KML.")


# --- Handler de Mensagem de Texto para Fluxos de Conversa ---

# Este handler atua como uma "máquina de estados" para gerenciar o fluxo após a conversão de um arquivo.
async def handle_mensagem(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    message = update.message or update.edited_message
    mensagem = message.text

    # ESTADO 1: O bot está esperando uma opção (0, 1 ou 2) após converter um arquivo.
    if context.user_data.get('MsgUser_ApplyPointTemplates'):
        if mensagem == "1":
            context.user_data['selected_flow'] = 1
            await context.bot.send_message(chat_id, "📌 Fluxo [1] selecionado!\n\nPor favor, informe o POP para continuar.")
        elif mensagem == "2":
            context.user_data['selected_flow'] = 2
            await context.bot.send_message(chat_id, "📌 Fluxo [2] selecionado!\n\nPor favor, informe o POP para continuar.")
        else:
            await message.reply_text("Comando 'convert' finalizado.")
            context.user_data.clear()
            return
        context.user_data.pop('MsgUser_ApplyPointTemplates')
        context.user_data['waiting_for_pop'] = True
        return

    # ESTADO 2: O bot está esperando um POP
    if context.user_data.get('waiting_for_pop'):
        pop_informado = mensagem.upper()
        flow = context.user_data.get('selected_flow')
        
        NomeCidade = await buscar_cidade_por_pop(pop_informado)
        if not NomeCidade:
            await message.reply_text("❌ POP não encontrado! Tente novamente ou digite 0 para sair.")
            return

        LinkDrive = await buscar_dir_drive()
        if "❌" in LinkDrive:
            await message.reply_text(LinkDrive)
            context.user_data.clear()
            return
        
        xlsx_file = context.user_data.get('xlsx_file')
        caminho_aux = os.path.join(LinkDrive, NomeCidade.replace("-", " "), "ARQUIVOS AUXILIARES")
        
        # Executa o fluxo de salvar o arquivo convertido
        await _run_blocking_io(EnviaArquivosDrive, caminho_aux, xlsx_file)
        await message.reply_text(f"✅ Arquivo '{xlsx_file}' salvo na pasta de arquivos auxiliares.")

        if flow == 2:
            # Lógica para inputar no template
            await message.reply_text("⚙️ Lógica para inputar os pontos no template ainda a ser implementada.")
            # ... aqui entraria a lógica de DE_KMZ_BASE_PARA_TEMPLATE, que também precisa ser não-bloqueante
        
        context.user_data.clear()
        await ExcluirArquivosporExtensao()

    # ESTADO 2 / FLUXO 1: O bot está esperando um POP para o fluxo 1.
    if context.user_data.get('waiting_for_pop_1'):
        NomeCidade = await buscar_cidade_por_pop(mensagem)
        if NomeCidade:
            caminho_do_arquivo = os.path.join(buscar_dir_drive(), NomeCidade.replace("-", " "), "ARQUIVOS AUXILIARES")
            # Recupera o nome do arquivo .xlsx que foi gerado e salvo no user_data.
            xlsx_file = context.user_data.get('xlsx_file')
            await EnviaArquivosDrive(caminho_do_arquivo, xlsx_file, chat_id, context)
            context.user_data.clear() # Limpa todos os dados da conversa.
            ExcluirArquivosporExtensao() # Exclui arquivos temporários locais.
        else:
            await update.message.reply_text("❌ POP não encontrado! Tente novamente ou digite 0 para sair.")
            
    # ESTADO 3 / FLUXO 2: O bot está esperando um POP para o fluxo 2.
    if context.user_data.get('waiting_for_pop_2'):
        PopInformado_user = mensagem.upper()
        NomeCidade = await buscar_cidade_por_pop(PopInformado_user)
        if NomeCidade:
            caminho_do_arquivo_aux = os.path.join(buscar_dir_drive(), NomeCidade.replace("-", " "), "ARQUIVOS AUXILIARES")
            DirTemplate = os.path.join(buscar_dir_drive(), NomeCidade.replace("-", " "), "CEP CTO")
            xlsx_file = context.user_data.get('xlsx_file')

            # Verifica se o arquivo de template existe no diretório.
            await VerificarTemplatemporPOP(DirTemplate, PopInformado_user, update)
            await EnviaArquivosDrive(caminho_do_arquivo_aux, xlsx_file, chat_id, context)

            ExcluirArquivosporExtensao()
            context.user_data.clear() # Limpa os dados da conversa ao finalizar.
        else:
            await update.message.reply_text("❌ POP não encontrado na lista de templates! O processo foi encerrado.")
            context.user_data.clear()
            ExcluirArquivosporExtensao()

async def mensagem_editada(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    # Se o update contiver uma mensagem editada, e essa mensagem tiver texto...
    if update.edited_message and update.edited_message.text:
        await handle_mensagem(update, context)

@check_permission
async def excluir_usuario(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """
    (Admin) Remove um utilizador pelo ID do Telegram ou pela matrícula.
    Uso: /excluir_usuario <id_do_telegram_ou_matricula>
    """
    message = update.message or update.edited_message
    if not message: return

    if not context.args or len(context.args) != 1:
        await message.reply_text("Uso correto: <code>/excluir_usuario &lt;ID ou Matrícula&gt;</code>", parse_mode=ParseMode.HTML)
        return

    identificador = context.args[0]
    if not identificador.isdigit():
        await message.reply_text("❌ O identificador (ID ou Matrícula) deve ser um número.")
        return
    
    id_ou_matricula = int(identificador)

    conexao_db = None
    try:
        conexao_db = await criar_conexao_db()
        if not conexao_db: raise ConnectionError("DB indisponível.")
        
        async with conexao_db.cursor() as cursor:
            # Tenta apagar onde o id_telegram OU a matricula correspondem ao identificador
            query = "DELETE FROM usuarios WHERE id_telegram = %s OR matricula = %s"
            rows_affected = await cursor.execute(query, (id_ou_matricula, id_ou_matricula))
        
        if rows_affected > 0:
            await message.reply_text(f"✅ Utilizador com ID/Matrícula <code>{id_ou_matricula}</code> foi removido com sucesso.", parse_mode=ParseMode.HTML)
            logger.info(f"Admin {update.effective_user.id} removeu o utilizador com ID/Matrícula {id_ou_matricula}.")
        else:
            await message.reply_text(f"⚠️ Nenhum utilizador encontrado com o ID ou Matrícula <code>{id_ou_matricula}</code>.", parse_mode=ParseMode.HTML)

    except Exception as e:
        logger.error(f"Erro em /excluir_usuario: {e}", exc_info=True)
        await message.reply_text("❌ Ocorreu um erro ao tentar remover o utilizador.")
    finally:
        if conexao_db:
            conexao_db.close()

@check_permission
async def novo_cargo(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """
    (Admin) Adiciona um novo cargo à tabela 'cargos'.
    Uso: /novo_cargo <NomeDoCargo>
    """
    message = update.message or update.edited_message
    if not message: return

    if not context.args:
        await message.reply_text("Uso correto: <code>/novo_cargo &lt;NomeDoCargo&gt;</code> (sem espaços)", parse_mode=ParseMode.HTML)
        return

    # Pega o nome do cargo e capitaliza a primeira letra
    nome_cargo = context.args[0].capitalize()
    conexao_db = None
    try:
        conexao_db = await criar_conexao_db()
        if not conexao_db: raise ConnectionError("DB indisponível.")
        
        async with conexao_db.cursor() as cursor:
            await cursor.execute("INSERT INTO cargos (nome_cargo) VALUES (%s)", (nome_cargo,))
        
        await message.reply_text(f"✅ Cargo '<b>{escape(nome_cargo)}</b>' criado com sucesso!", parse_mode=ParseMode.HTML)
        logger.info(f"Admin {update.effective_user.id} criou o novo cargo: {nome_cargo}.")

    except aiomysql.IntegrityError:
        # Este erro ocorre se o cargo já existir (devido à restrição UNIQUE)
        await message.reply_text(f"⚠️ O cargo '<b>{escape(nome_cargo)}</b>' já existe.", parse_mode=ParseMode.HTML)
    except Exception as e:
        logger.error(f"Erro em /novo_cargo: {e}", exc_info=True)
        await message.reply_text("❌ Ocorreu um erro ao criar o novo cargo.")
    finally:
        if conexao_db:
            conexao_db.close()

@check_permission
async def excluir_cargo(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """
    (Admin) Remove um cargo da tabela 'cargos'.
    Uso: /excluir_cargo <NomeDoCargo>
    """
    message = update.message or update.edited_message
    if not message: return

    if not context.args:
        await message.reply_text("Uso correto: <code>/excluir_cargo &lt;NomeDoCargo&gt;</code>", parse_mode=ParseMode.HTML)
        return

    nome_cargo = context.args[0].capitalize()
    conexao_db = None
    try:
        conexao_db = await criar_conexao_db()
        if not conexao_db: raise ConnectionError("DB indisponível.")
        
        async with conexao_db.cursor() as cursor:
            # ON DELETE CASCADE irá remover as permissões associadas
            rows_affected = await cursor.execute("DELETE FROM cargos WHERE nome_cargo = %s", (nome_cargo,))
        
        if rows_affected > 0:
            await message.reply_text(f"✅ Cargo '<b>{escape(nome_cargo)}</b>' e todas as suas permissões foram removidos.", parse_mode=ParseMode.HTML)
            logger.info(f"Admin {update.effective_user.id} removeu o cargo {nome_cargo}.")
        else:
            await message.reply_text(f"⚠️ Nenhum cargo encontrado com o nome '<b>{escape(nome_cargo)}</b>'.", parse_mode=ParseMode.HTML)

    except aiomysql.IntegrityError:
        await message.reply_text(f"❌ Não é possível remover o cargo '<b>{escape(nome_cargo)}</b>' pois ele ainda está em uso por algum utilizador.", parse_mode=ParseMode.HTML)
    except Exception as e:
        logger.error(f"Erro em /excluir_cargo: {e}", exc_info=True)
        await message.reply_text("❌ Ocorreu um erro ao tentar remover o cargo.")
    finally:
        if conexao_db:
            conexao_db.close()


# --- Funções de Gestão de Comandos e Permissões ---

@check_permission
async def novo_comando(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """
    (Admin) Adiciona um novo comando e concede permissão automática ao 'Administrador'.
    Uso: /novo_comando <nome_do_comando>
    """
    message = update.message or update.edited_message
    if not message: return

    if not context.args:
        await message.reply_text("Uso correto: <code>/novo_comando &lt;nome_do_comando&gt;</code>", parse_mode=ParseMode.HTML)
        return

    nome_comando = context.args[0].lower() # Comandos são minúsculos por convenção
    conexao_db = None
    try:
        conexao_db = await criar_conexao_db()
        if not conexao_db: raise ConnectionError("DB indisponível.")
        
        async with conexao_db.cursor() as cursor:
            # 1. Insere o novo comando. Falhará se já existir (IntegrityError).
            await cursor.execute("INSERT INTO comandos (nome_comando) VALUES (%s)", (nome_comando,))
            
            # 2. Concede permissão ao 'Administrador' para o novo comando.
            query_permission = """
                INSERT INTO permissoes (cargo_id, comando_id)
                VALUES (
                    (SELECT id FROM cargos WHERE nome_cargo = 'Administrador'),
                    (SELECT id FROM comandos WHERE nome_comando = %s)
                )
            """
            await cursor.execute(query_permission, (nome_comando,))
        
        await message.reply_text(
            f"✅ Comando <code>/{nome_comando}</code> adicionado com sucesso!\n"
            f"Permissão automática concedida ao cargo <b>Administrador</b>.",
            parse_mode=ParseMode.HTML
        )
        logger.info(f"Admin {update.effective_user.id} adicionou o novo comando '{nome_comando}' com permissão de admin.")

    except aiomysql.IntegrityError:
        await message.reply_text(f"⚠️ O comando <code>/{nome_comando}</code> já existe no sistema.", parse_mode=ParseMode.HTML)
    except Exception as e:
        logger.error(f"Erro em /novo_comando: {e}", exc_info=True)
        await message.reply_text("❌ Ocorreu um erro ao adicionar o novo comando.")
    finally:
        if conexao_db:
            conexao_db.close()

@check_permission
async def excluir_comando(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """
    (Admin) Remove um comando da tabela 'comandos'.
    Uso: /excluir_comando <nome_do_comando>
    """
    message = update.message or update.edited_message
    if not message: return

    if not context.args:
        await message.reply_text("Uso correto: <code>/excluir_comando &lt;nome_do_comando&gt;</code>", parse_mode=ParseMode.HTML)
        return

    nome_comando = context.args[0].lower()
    conexao_db = None
    try:
        conexao_db = await criar_conexao_db()
        if not conexao_db: raise ConnectionError("DB indisponível.")
        
        async with conexao_db.cursor() as cursor:
            # ON DELETE CASCADE irá remover as permissões associadas
            rows_affected = await cursor.execute("DELETE FROM comandos WHERE nome_comando = %s", (nome_comando,))
        
        if rows_affected > 0:
            await message.reply_text(f"✅ Comando <code>/{nome_comando}</code> e todas as suas permissões foram removidos.", parse_mode=ParseMode.HTML)
            logger.info(f"Admin {update.effective_user.id} removeu o comando {nome_comando}.")
        else:
            await message.reply_text(f"⚠️ Nenhum comando encontrado com o nome '<code>/{nome_comando}</code>'.", parse_mode=ParseMode.HTML)

    except Exception as e:
        logger.error(f"Erro em /excluir_comando: {e}", exc_info=True)
        await message.reply_text("❌ Ocorreu um erro ao tentar remover o comando.")
    finally:
        if conexao_db:
            conexao_db.close()

@check_permission
async def limpar_convites(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """
    (Admin) Limpa todos os registos da tabela de cadastros pendentes.
    Uso: /limpar_convites
    """
    message = update.message or update.edited_message
    if not message: return

    conexao_db = None
    try:
        conexao_db = await criar_conexao_db()
        if not conexao_db: raise ConnectionError("DB indisponível.")
        
        async with conexao_db.cursor() as cursor:
            # Executa o comando DELETE e guarda o número de linhas removidas
            rows_deleted = await cursor.execute("DELETE FROM cadastros_pendentes")
        
        await message.reply_text(f"🧹 Limpeza concluída! {rows_deleted} convite(s) pendente(s) foram removidos.")
        logger.info(f"Admin {update.effective_user.id} limpou a tabela de convites pendentes.")

    except Exception as e:
        logger.error(f"Erro em /limpar_convites: {e}", exc_info=True)
        await message.reply_text("❌ Ocorreu um erro ao limpar os convites pendentes.")
    finally:
        if conexao_db:
            conexao_db.close()

@check_permission
async def adicionar_permissao(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """
    (Admin) Concede a um cargo permissão para usar um comando.
    Uso: /adicionar_permissao <Cargo> <nome_do_comando>
    """
    message = update.message or update.edited_message
    if not message: return

    if len(context.args) != 2:
        await message.reply_text("Uso: <code>/adicionar_permissao &lt;Cargo&gt; &lt;comando&gt;</code>\nEx: /adicionar_permissao Tecnico ctos", parse_mode=ParseMode.HTML)
        return

    cargo, comando = context.args
    cargo = cargo.capitalize()
    comando = comando.lower() 
    conexao_db = None
    try:
        conexao_db = await criar_conexao_db()
        if not conexao_db: raise ConnectionError("DB indisponível.")
        
        async with conexao_db.cursor() as cursor:
            query = """
                INSERT INTO permissoes (cargo_id, comando_id)
                VALUES (
                    (SELECT id FROM cargos WHERE nome_cargo = %s),
                    (SELECT id FROM comandos WHERE nome_comando = %s)
                )
            """
            await cursor.execute(query, (cargo, comando))
        
        await message.reply_text(f"✅ Permissão concedida! O cargo <b>{escape(cargo)}</b> agora pode usar o comando <code>/{comando}</code>.", parse_mode=ParseMode.HTML)
        logger.info(f"Admin {update.effective_user.id} concedeu a permissão /{comando} para o cargo {cargo}.")

    except aiomysql.IntegrityError:
        await message.reply_text(f"❌ Falha ao adicionar permissão. Verifique se o cargo '<b>{escape(cargo)}</b>' e o comando '<code>/{comando}</code>' existem e se a permissão já não foi concedida.", parse_mode=ParseMode.HTML)
    except Exception as e:
        logger.error(f"Erro em /adicionar_permissao: {e}", exc_info=True)
        await message.reply_text("❌ Ocorreu um erro ao adicionar a permissão.")
    finally:
        if conexao_db:
            conexao_db.close()
            
@check_permission
async def revogar_permissao(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """
    (Admin) Revoga de um cargo a permissão para usar um comando.
    Uso: /revogar_permissao <Cargo> <nome_do_comando>
    """
    message = update.message or update.edited_message
    if not message: return

    if len(context.args) != 2:
        await message.reply_text("Uso: <code>/revogar_permissao &lt;Cargo&gt; &lt;comando&gt;</code>\nEx: /revogar_permissao Tecnico ctos", parse_mode=ParseMode.HTML)
        return

    cargo, comando = context.args
    cargo = cargo.capitalize()
    comando = comando.lower()
    conexao_db = None
    try:
        conexao_db = await criar_conexao_db()
        if not conexao_db: raise ConnectionError("DB indisponível.")
        
        async with conexao_db.cursor() as cursor:
            query = """
                DELETE FROM permissoes
                WHERE cargo_id = (SELECT id FROM cargos WHERE nome_cargo = %s)
                AND comando_id = (SELECT id FROM comandos WHERE nome_comando = %s)
            """
            rows_affected = await cursor.execute(query, (cargo, comando))

        if rows_affected > 0:
            await message.reply_text(f"✅ Permissão revogada! O cargo <b>{escape(cargo)}</b> não pode mais usar o comando <code>/{comando}</code>.", parse_mode=ParseMode.HTML)
            logger.info(f"Admin {update.effective_user.id} revogou a permissão /{comando} do cargo {cargo}.")
        else:
            await message.reply_text("⚠️ Nenhuma permissão correspondente encontrada para ser revogada. Verifique os nomes do cargo e do comando.")

    except Exception as e:
        logger.error(f"Erro em /revogar_permissao: {e}", exc_info=True)
        await message.reply_text("❌ Ocorreu um erro ao revogar a permissão.")
    finally:
        if conexao_db:
            conexao_db.close()

async def verificar_inativos(context: ContextTypes.DEFAULT_TYPE) -> None:
    logger.info("A executar job diário: Verificação de utilizadores inativos...")
    conexao_db = None
    try:
        conexao_db = await criar_conexao_db()
        if not conexao_db:
            logger.error("Job 'verificar_inativos': Não foi possível conectar ao DB.")
            return

        async with conexao_db.cursor() as cursor:
            # Query para encontrar e atualizar utilizadores ativos com mais de 30 dias de inatividade
            query = """
                UPDATE usuarios
                SET esta_ativo = FALSE
                WHERE esta_ativo = TRUE AND ultima_interacao < DATE_SUB(NOW(), INTERVAL 30 DAY);
            """
            rows_affected = await cursor.execute(query)
        
        if rows_affected > 0:
            logger.info(f"Job 'verificar_inativos': {rows_affected} utilizador(es) foram marcados como inativos.")
        else:
            logger.info("Job 'verificar_inativos': Nenhum utilizador inativo encontrado.")

    except Exception as e:
        logger.error(f"Job 'verificar_inativos': Falha ao executar a tarefa. Erro: {e}")
    finally:
        if conexao_db:
            conexao_db.close()

#Comando /reativar_usuario
@check_permission
async def reativar_usuario(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """
    (Admin) Reativa um utilizador que foi marcado como inativo.
    Uso: /reativar_usuario <ID ou Matrícula>
    """
    message = update.message or update.edited_message
    if not message: return

    if not context.args or len(context.args) != 1:
        # --- CORREÇÃO APLICADA AQUI ---
        # Substituído '<' e '>' por '&lt;' e '&gt;' para evitar erro de parse HTML
        await message.reply_text("Uso correto: <code>/reativar_usuario &lt;ID ou Matrícula&gt;</code>", parse_mode=ParseMode.HTML)
        return

    identificador = context.args[0]
    if not identificador.isdigit():
        await message.reply_text("❌ O identificador (ID ou Matrícula) deve ser um número.")
        return
    
    id_ou_matricula = int(identificador)

    conexao_db = None
    try:
        conexao_db = await criar_conexao_db()
        if not conexao_db: raise ConnectionError("DB indisponível.")
        
        async with conexao_db.cursor() as cursor:
            # Reativa a conta e atualiza a data de interação para o momento atual
            query = """
                UPDATE usuarios 
                SET esta_ativo = TRUE, ultima_interacao = NOW() 
                WHERE id_telegram = %s OR matricula = %s
            """
            rows_affected = await cursor.execute(query, (id_ou_matricula, id_ou_matricula))
        
        if rows_affected > 0:
            await message.reply_text(f"✅ Utilizador com ID/Matrícula <code>{id_ou_matricula}</code> foi reativado com sucesso.", parse_mode=ParseMode.HTML)
            logger.info(f"Admin {update.effective_user.id} reativou o utilizador com ID/Matrícula {id_ou_matricula}.")
        else:
            await message.reply_text(f"⚠️ Nenhum utilizador inativo encontrado com o ID ou Matrícula <code>{id_ou_matricula}</code>.", parse_mode=ParseMode.HTML)

    except Exception as e:
        logger.error(f"Erro em /reativar_usuario: {e}", exc_info=True)
        await message.reply_text("❌ Ocorreu um erro ao tentar reativar o utilizador.")
    finally:
        if conexao_db:
            conexao_db.close()

# --- Função Principal de Execução do Bot ---

def main() -> None:
 
    try:
        app = ApplicationBuilder().token(BOT_TOKEN).connect_timeout(10).read_timeout(10).job_queue(JobQueue()).build()
        
        # --- Registro de Handlers ---

        # Handler de Erro.
        app.add_error_handler(error_handler)

        # Handler de Reconexão:
        app.add_handler(MessageHandler(filters.ALL, check_reconnection), group=-1)


        # Handler de Conversa.
        conv_handler_novo_usuario = ConversationHandler(
            entry_points=[CommandHandler("novo_usuario", novo_usuario)],
            states={
                RECEBER_MATRICULA: [MessageHandler(filters.TEXT & ~filters.COMMAND, receber_matricula)],
                RECEBER_NOME: [MessageHandler(filters.TEXT & ~filters.COMMAND, receber_nome)],
            },
            fallbacks=[CommandHandler("cancelar", cancelar_cadastro)],
            per_message=False
        )
        app.add_handler(conv_handler_novo_usuario)

        # Mapa de Comandos.
        app.add_handler(CommandHandler("start", ajuda))
        app.add_handler(CommandHandler("ajuda", ajuda))
        app.add_handler(CommandHandler("ctos", ctos))
        app.add_handler(CommandHandler("novaCTO", novacto))
        app.add_handler(CommandHandler("atividades", atividades))
        app.add_handler(CommandHandler("checar", checar))
        app.add_handler(CommandHandler("localizar", localizar_cto))
        app.add_handler(CommandHandler("ExibirCidade", exibircidade))
        app.add_handler(CommandHandler("input", input))
        app.add_handler(CommandHandler("insert", insert))
        app.add_handler(CommandHandler("listarIDs", listarids))
        app.add_handler(CommandHandler("convert", convert))
        app.add_handler(CommandHandler("gerarkmzatualizado", gerarkmzatualizado))
        app.add_handler(CommandHandler("baixarkmz", baixarkmz))
        app.add_handler(CommandHandler("Id", id))
        app.add_handler(CommandHandler("Info", info))
        # Comandos de administração
        app.add_handler(CommandHandler("AjudaAdm", ajudaadm))
        app.add_handler(CommandHandler("CWH", CWH))
        app.add_handler(CommandHandler("AdcionarTemplate", adicionartemplate))
        app.add_handler(CommandHandler("ExcluirTemplate", excluirtemplate))
        app.add_handler(CommandHandler("configdrive", configdrive))
        app.add_handler(CommandHandler("listar_admins", listar_admins))
        app.add_handler(CommandHandler("cadastrar", cadastrar))
        app.add_handler(CommandHandler("excluir_usuario", excluir_usuario))
        app.add_handler(CommandHandler("novo_cargo", novo_cargo))
        app.add_handler(CommandHandler("excluir_cargo", excluir_cargo))
        app.add_handler(CommandHandler("adicionar_permissao", adicionar_permissao))
        app.add_handler(CommandHandler("revogar_permissao", revogar_permissao))
        app.add_handler(CommandHandler("novo_comando", novo_comando))
        app.add_handler(CommandHandler("excluir_comando", excluir_comando))
        app.add_handler(CommandHandler("limpar_convites", limpar_convites))
        app.add_handler(CommandHandler("reativar_usuario", reativar_usuario))

        
        # Handlers de Mensagem.
        # Handler para qualquer mensagem de localização.
        map_pattern = r'(maps\.google\.com|goo\.gl/maps|waze\.com|@?(-?\d+\.\d+)[, ](-?\d+\.\d+))'
        app.add_handler(MessageHandler(filters.LOCATION | filters.Regex(map_pattern),unified_location_handler))        
        # Handler para qualquer tipo de documento enviado.
        app.add_handler(MessageHandler(filters.Document.ALL, handle_arquivo))
        # Handler para qualquer mensagem de texto que NÃO seja um comando.
        app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_mensagem))
        # Handler que escuta por edições de mensagens de texto e envia para a mesma função 'handle_mensagem'.
        app.add_handler(TypeHandler(Update, mensagem_editada))


        # --- Agendamento de Tarefas ---
        fuso_horario_sp = pytz.timezone('America/Sao_Paulo')
        
        #atualizar_admins_fallback
        horario = dt_time(hour=3, minute=0, second=0, tzinfo=fuso_horario_sp)
        logger.info(f"Agendando atualizar_admins_fallback para as {horario.strftime('%H:%M:%S %Z')}")
        # Agenda a função 'atualizar_admins_fallback' para rodar diariamente no horário definido.
        app.job_queue.run_daily(atualizar_admins_fallback, time=horario, name="Atualização lista Admins")
        
        #verificar_inativos
        horario_inativos = dt_time(hour=4, minute=0, second=0, tzinfo=fuso_horario_sp)
        logger.info(f"Agendando verificar_inativos para as {horario_inativos.strftime('%H:%M:%S %Z')}")
        # Agenda a função 'verificar_inativos' para rodar diariamente no horário definido.
        app.job_queue.run_daily(verificar_inativos,time=horario_inativos,name="Verificação de Utilizadores Inativos")

        logger.info("Automação está rodando...")
        app.run_polling()

    except Exception as e:
        logger.critical(f"Erro fatal ao iniciar ou executar o bot: {e}", exc_info=True)


if __name__ == "__main__":
    main()