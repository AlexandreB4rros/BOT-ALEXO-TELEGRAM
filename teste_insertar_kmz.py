import openpyxl

def copiar_dados_xlsx(arquivo_origem, arquivo_destino):
    """Copia os dados das colunas A, B e C do arquivo de origem para a planilha 'KMZ' no arquivo de destino."""
    try:
        wb_origem = openpyxl.load_workbook(arquivo_origem)
        sheet_origem = wb_origem.active  # Usa a primeira planilha
        
        wb_destino = openpyxl.load_workbook(arquivo_destino)

        if "KMZ" not in wb_destino.sheetnames:
            sheet_destino = wb_destino.create_sheet("KMZ")  # Cria a planilha se não existir
        else:
            sheet_destino = wb_destino["KMZ"]

        for row_idx, row in enumerate(sheet_origem.iter_rows(min_row=3, values_only=True), start=3):
            sheet_destino[f"A{row_idx}"] = row[0]  # Coluna A (Nome do ponto)
            sheet_destino[f"B{row_idx}"] = row[1]  # Coluna B (Latitude)
            sheet_destino[f"C{row_idx}"] = row[2]  # Coluna C (Longitude)

        wb_destino.save(arquivo_destino)

        print("✅ Dados copiados com sucesso para a planilha 'KMZ'!")

    except Exception as e:
        print(f"❌ Erro ao copiar os dados: {e}")

# Exemplo de uso
copiar_dados_xlsx("arquv/TTI - TATUI CENTRAL WAVE (6).xlsx", "TEMPLATE REDES IPERÓ.xlsx")
