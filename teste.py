import openpyxl
import simplekml

def converter_planilha_template_para_kml(caminho_xlsx, nome_planilha, caminho_kml):
    """Essa função cria o kml acessando o template na planilha KMZ e gera o KMZ BASE"""
    try:
        # Abre o arquivo Excel
        workbook = openpyxl.load_workbook(caminho_xlsx)
        
        # Verifica se a planilha existe
        if nome_planilha not in workbook.sheetnames:
            print(f"❌ A planilha '{nome_planilha}' não foi encontrada.")
            print(f"📄 Planilhas disponíveis: {workbook.sheetnames}")
            return
        
        # Seleciona a planilha
        sheet = workbook[nome_planilha]

        # Cria um objeto KML
        kml = simplekml.Kml()

        # Lê os dados a partir da terceira linha (ignorando cabeçalho)
        for row in sheet.iter_rows(min_row=3, values_only=True):
            nome = row[0]  # Coluna A (Nome do ponto)
            lat = row[1]   # Coluna B (Latitude)
            lon = row[2]   # Coluna C (Longitude)

            # Verifica se os valores são válidos
            if nome and lat and lon:
                pnt = kml.newpoint(name=str(nome), coords=[(lon, lat)])  # (Longitude, Latitude)

                # Define o ícone do ponto
                pnt.style.iconstyle.icon.href = icone_url
                pnt.style.iconstyle.scale = 1.5  # Tamanho do ícone

        # Salva o arquivo KML
        kml.save(caminho_kml)
        print(f"✅ Arquivo KML gerado com sucesso: {caminho_kml}")

        # Fecha o arquivo Excel
        workbook.close()
    
    except Exception as e:
        print(f"❌ Erro ao processar o arquivo: {e}")

# Exemplo de uso
caminho_xlsx = "TEMPLATE REDES IPERÓ.xlsx"
nome_da_planilha = "KMZ"

NOMECIDADE = "GEORGE OETTERER"
POP = "IEG"

caminho_kml = f"{POP} - {NOMECIDADE} - KMZ BASE.kml"
icone_url = "http://maps.google.com/mapfiles/kml/shapes/placemark_circle.png"

converter_planilha_template_para_kml(caminho_xlsx, nome_da_planilha, caminho_kml)
